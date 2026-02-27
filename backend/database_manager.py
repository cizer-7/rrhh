import mysql.connector
from mysql.connector import Error
from mysql.connector import pooling
from typing import Dict, List, Any, Optional
import logging
import hashlib
from database_exports import DatabaseManagerExportsMixin
from datetime import datetime
import json
import os
from decimal import Decimal
from openpyxl.worksheet.worksheet import Worksheet

class DatabaseManager(DatabaseManagerExportsMixin):
    def __init__(self, host: str, database: str, user: str, password: str, port: int = 3307, ssl_disabled: Optional[bool] = None):
        self.host = host
        self.database = database
        self.user = user
        self.password = password
        self.port = port
        self.ssl_disabled = ssl_disabled
        self.connection = None
        self._pool = None
        self.logger = logging.getLogger(__name__)

    def _ensure_carry_over_table(self) -> None:
        try:
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT COUNT(*) AS cnt
                FROM INFORMATION_SCHEMA.TABLES
                WHERE TABLE_SCHEMA = %s
                  AND TABLE_NAME = 't010_carry_over'
                """,
                (self.database,),
            )
            rows = cursor.fetchall() or []
            cursor.close()
            cnt = int(rows[0].get('cnt', 0)) if rows else 0
            if cnt > 0:
                return

            cursor = self.connection.cursor()
            cursor.execute(
                """
                CREATE TABLE IF NOT EXISTS t010_carry_over (
                    id_carry_over INT AUTO_INCREMENT PRIMARY KEY,
                    id_empleado INT NOT NULL,
                    source_anio INT NOT NULL,
                    source_mes INT NOT NULL,
                    apply_anio INT NOT NULL,
                    apply_mes INT NOT NULL,
                    concept VARCHAR(50) NOT NULL,
                    amount DECIMAL(12,2) NOT NULL DEFAULT 0.00,
                    created_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP,
                    updated_at TIMESTAMP DEFAULT CURRENT_TIMESTAMP ON UPDATE CURRENT_TIMESTAMP,
                    INDEX idx_t010_apply (apply_anio, apply_mes),
                    INDEX idx_t010_source (source_anio, source_mes),
                    INDEX idx_t010_empleado (id_empleado),
                    FOREIGN KEY (id_empleado) REFERENCES t001_empleados(id_empleado) ON DELETE CASCADE
                )
                """
            )
            cursor.close()
        except Exception as e:
            self.logger.error(f"Fehler beim Sicherstellen der t010_carry_over Tabelle: {e}")

    def _ensure_employee_hire_date_column(self) -> None:
        """Ensures t001_empleados has fecha_alta column (hire date).
        This is a lightweight migration to keep existing installations working.
        """
        try:
            check_query = """
            SELECT COUNT(*) AS cnt
            FROM INFORMATION_SCHEMA.COLUMNS
            WHERE TABLE_SCHEMA = %s
              AND TABLE_NAME = 't001_empleados'
              AND COLUMN_NAME = 'fecha_alta'
            """
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute(check_query, (self.database,))
            rows = cursor.fetchall()
            cursor.close()
            
            cnt = int(rows[0].get('cnt', 0)) if rows else 0
            if cnt > 0:
                return
            alter_query = "ALTER TABLE t001_empleados ADD COLUMN fecha_alta DATE NULL"
            cursor = self.connection.cursor()
            cursor.execute(alter_query)
            cursor.close()
        except Exception as e:
            self.logger.error(f"Fehler beim Sicherstellen von fecha_alta Spalte: {e}")

    def _ensure_employee_irpf_columns(self) -> None:
        try:
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT COLUMN_NAME
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = %s
                  AND TABLE_NAME = 't001_empleados'
                  AND COLUMN_NAME IN ('declaracion', 'dni')
                """,
                (self.database,),
            )
            rows = cursor.fetchall() or []
            cursor.close()

            existing = {str(r.get('COLUMN_NAME', '')).strip().lower() for r in rows}

            if 'declaracion' not in existing:
                cursor = self.connection.cursor()
                cursor.execute(
                    "ALTER TABLE t001_empleados ADD COLUMN declaracion VARCHAR(20) NULL DEFAULT NULL"
                )
                cursor.close()

            if 'dni' not in existing:
                cursor = self.connection.cursor()
                cursor.execute("ALTER TABLE t001_empleados ADD COLUMN dni VARCHAR(50) NULL DEFAULT NULL")
                cursor.close()
        except Exception as e:
            self.logger.error(f"Fehler beim Sicherstellen von declaracion/dni Spalten: {e}")

    def _ensure_deducciones_gasolina_column(self) -> None:
        """Ensures t004_deducciones_mensuales has a single 'gasolina' column.

        Existing data in those legacy columns is intentionally ignored.
        """
        try:
            cursor = self.connection.cursor(dictionary=True)
            cursor.execute(
                """
                SELECT COLUMN_NAME
                FROM INFORMATION_SCHEMA.COLUMNS
                WHERE TABLE_SCHEMA = %s
                  AND TABLE_NAME = 't004_deducciones_mensuales'
                  AND COLUMN_NAME IN ('gasolina', 'gasolina_arval', 'gasolina_ald')
                """,
                (self.database,),
            )
            rows = cursor.fetchall() or []
            cursor.close()

            existing = {str(r.get('COLUMN_NAME', '')).strip().lower() for r in rows}
            if 'gasolina' not in existing:
                cursor = self.connection.cursor()
                cursor.execute("ALTER TABLE t004_deducciones_mensuales ADD COLUMN gasolina DECIMAL(10,2) DEFAULT 0.00")
                cursor.close()

            drop_cols = []
            if 'gasolina_arval' in existing:
                drop_cols.append('gasolina_arval')
            if 'gasolina_ald' in existing:
                drop_cols.append('gasolina_ald')

            if drop_cols:
                cursor = self.connection.cursor()
                cursor.execute(
                    f"ALTER TABLE t004_deducciones_mensuales {', '.join([f'DROP COLUMN {c}' for c in drop_cols])}"
                )
                cursor.close()
        except Exception as e:
            self.logger.error(f"Fehler beim Sicherstellen von gasolina Spalte: {e}")

    def insert_registro_procesamiento(
        self,
        usuario_login: str,
        accion: str,
        objeto: str = None,
        id_empleado: int = None,
        anio: int = None,
        mes: int = None,
        detalles: Dict[str, Any] = None,
    ) -> bool:
        try:
            if not usuario_login or not accion:
                return False
            detalles_json = None
            if detalles is not None:
                try:
                    detalles_json = json.dumps(detalles, ensure_ascii=False)
                except Exception:
                    detalles_json = None
            query = """
            INSERT INTO t007_registro_procesamiento (usuario_login, id_empleado, anio, mes, accion, objeto, detalles)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            """
            return self.execute_update(
                query,
                (
                    usuario_login,
                    id_empleado,
                    anio,
                    mes,
                    accion,
                    objeto,
                    detalles_json,
                ),
            )
        except Exception as e:
            self.logger.error(f"Fehler beim Schreiben in t007_registro_procesamiento: {e}")
            return False

    def get_registro_procesamiento(
        self,
        id_empleado: int,
        anio: int = None,
        mes: int = None,
        limit: int = 200,
    ) -> List[Dict]:
        try:
            if id_empleado is None:
                return []
            limit_value = 200
            try:
                if isinstance(limit, int):
                    limit_value = max(1, min(limit, 1000))
            except Exception:
                limit_value = 200
            query = """
            SELECT
                l.id_registro,
                l.fecha,
                l.usuario_login,
                u.nombre_completo,
                l.id_empleado,
                l.anio,
                l.mes,
                l.accion,
                l.objeto,
                l.detalles
            FROM t007_registro_procesamiento l
            LEFT JOIN t005_usuarios u ON u.nombre_usuario = l.usuario_login
            WHERE l.id_empleado = %s
            """
            params: List[Any] = [id_empleado]
            if anio is not None:
                query += " AND l.anio = %s"
                params.append(anio)
            if mes is not None:
                query += " AND l.mes = %s"
                params.append(mes)
            query += " ORDER BY l.fecha DESC, l.id_registro DESC LIMIT %s"
            params.append(limit_value)
            rows = self.execute_query(query, tuple(params))
            for r in rows:
                if isinstance(r.get('detalles'), str):
                    try:
                        r['detalles'] = json.loads(r['detalles'])
                    except Exception:
                        pass
            return rows
        except Exception as e:
            self.logger.error(f"Fehler beim Lesen von t007_registro_procesamiento: {e}")
            return []

    def get_global_registro_procesamiento(
        self,
        id_empleado: int = None,
        anio: int = None,
        mes: int = None,
        limit: int = 200,
    ) -> List[Dict]:
        try:
            limit_value = 200
            try:
                if isinstance(limit, int):
                    limit_value = max(1, min(limit, 1000))
            except Exception:
                limit_value = 200
            query = """
            SELECT
                l.id_registro,
                l.fecha,
                l.usuario_login,
                u.nombre_completo,
                l.id_empleado,
                e.nombre as empleado_nombre,
                e.apellido as empleado_apellido,
                l.anio,
                l.mes,
                l.accion,
                l.objeto,
                l.detalles
            FROM t007_registro_procesamiento l
            LEFT JOIN t005_usuarios u ON u.nombre_usuario = l.usuario_login
            LEFT JOIN t001_empleados e ON e.id_empleado = l.id_empleado
            WHERE 1=1
            """
            params: List[Any] = []
            if id_empleado is not None:
                query += " AND l.id_empleado = %s"
                params.append(id_empleado)
            if anio is not None:
                query += " AND l.anio = %s"
                params.append(anio)
            if mes is not None:
                query += " AND l.mes = %s"
                params.append(mes)
            query += " ORDER BY l.fecha DESC, l.id_registro DESC LIMIT %s"
            params.append(limit_value)
            rows = self.execute_query(query, tuple(params))
            for r in rows:
                if isinstance(r.get('detalles'), str):
                    try:
                        r['detalles'] = json.loads(r['detalles'])
                    except Exception:
                        pass
            return rows
        except Exception as e:
            self.logger.error(f"Fehler beim Lesen von global t007_registro_procesamiento: {e}")
            return []

    def create_change_details(self, old_data: dict = None, new_data: dict = None, changed_fields: list = None) -> dict:
        """Erstellt Details mit Vorher/Nachher Werten für geänderte Felder"""
        details = {}
        
        if old_data and new_data:
            # Für Updates - nur tatsächlich geänderte Felder anzeigen
            for field in new_data.keys():
                if field in old_data or field in new_data:
                    old_value = old_data.get(field)
                    new_value = new_data.get(field)
                    
                    # Convert Decimal to float for consistent comparison
                    if isinstance(old_value, Decimal):
                        old_value = float(old_value)
                    if isinstance(new_value, Decimal):
                        new_value = float(new_value)
                    
                    # Convert datetime objects to string for consistent comparison
                    if isinstance(old_value, datetime):
                        old_value = old_value.isoformat()
                    if isinstance(new_value, datetime):
                        new_value = new_value.isoformat()
                    
                    # Nur hinzufügen, wenn sich der Wert tatsächlich geändert hat
                    if old_value != new_value:
                        details[field] = {
                            'old': old_value,
                            'new': new_value
                        }
        elif new_data:
            # Für neue Datensätze (create) - alle Felder als neu markieren
            for key, value in new_data.items():
                if key not in ['id_empleado', 'created_at', 'updated_at']:
                    details[key] = {
                        'old': None,
                        'new': value
                    }
        
        return details

    def _normalize_employee_category(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        if isinstance(value, str):
            v = value.strip()
            return v or None
        v = str(value).strip()
        return v or None

    def _is_valid_employee_category(self, value: Any) -> bool:
        v = self._normalize_employee_category(value)
        return v is None or v in {"Techniker", "Office"}

    def _create_connection(self):
        if self._pool is None:
            pool_kwargs: Dict[str, Any] = {
                "pool_name": "nomina_pool",
                "pool_size": 10,
                "pool_reset_session": True,
                "host": self.host,
                "database": self.database,
                "user": self.user,
                password=self.password,
                port=self.port,
                connection_timeout=10,
            }
            if self.ssl_disabled is not None:
                pool_kwargs["ssl_disabled"] = bool(self.ssl_disabled)
            self._pool = pooling.MySQLConnectionPool(**pool_kwargs)
        return self._pool.get_connection()

    def _is_transient_connection_error(self, err: Exception) -> bool:
        if not isinstance(err, Error):
            return False
        errno = getattr(err, "errno", None)
        if errno in (2006, 2013, 2055):
            return True
        msg = str(err)
        return (
            "Lost connection" in msg
            or "MySQL Connection not available" in msg
            or "EOF occurred" in msg
        )

    def connect(self):
        try:
            self.connection = self._create_connection()
            if self.connection.is_connected():
                self.logger.info(f"Erfolgreich verbunden mit MySQL Datenbank {self.database}")
                self._ensure_employee_hire_date_column()
                self._ensure_employee_irpf_columns()
                self._ensure_deducciones_gasolina_column()
                self._ensure_carry_over_table()
                return True
        except Error as e:
            self.logger.error(f"Fehler bei der Verbindung zur Datenbank: {e}")
            return False
        return False

    def _next_year_month(self, year: int, month: int) -> (int, int):
        y = int(year)
        m = int(month)
        if m >= 12:
            return y + 1, 1
        return y, m + 1

    def create_carry_over_batch(
        self,
        employee_id: int,
        source_year: int,
        source_month: int,
        items: List[Dict[str, Any]],
        defer_concepts: List[str],
    ) -> bool:
        """Creates carry over entries for the month after source_year/source_month.

        Behavior:
        - The user enters carry over values for the selected source month.
        - These values are applied in the following month.
        - Only the current user input is used; previous months' carry overs are not carried forward.

        Notes:
        - The save operation is idempotent for a given (employee, source_year, source_month):
          it replaces previously stored carry overs for that source month.
        - defer_concepts is accepted for backward compatibility but is ignored.
        """
        connection = None
        cursor = None
        try:
            if employee_id is None:
                return False
            source_year_i = int(source_year)
            source_month_i = int(source_month)
            apply_year, apply_month = self._next_year_month(source_year_i, source_month_i)

            # Don't collect carry overs from previous months anymore
            # Only use the current user input for carry over
            carry_forward: Dict[str, float] = {}

            connection = self._create_connection()
            cursor = connection.cursor()

            # Replace existing carry overs for this source month
            cursor.execute(
                """
                DELETE FROM t010_carry_over
                WHERE id_empleado = %s
                  AND source_anio = %s
                  AND source_mes = %s
                """,
                (int(employee_id), source_year_i, source_month_i),
            )

            # Insert new entries
            insert_query = """
            INSERT INTO t010_carry_over
                (id_empleado, source_anio, source_mes, apply_anio, apply_mes, concept, amount)
            VALUES
                (%s, %s, %s, %s, %s, %s, %s)
            """

            # Only use user input for carry over (no carry forward from previous months)
            combined: Dict[str, float] = {}
            for it in items or []:
                concept = str(it.get('concept', '')).strip()
                if not concept:
                    continue
                amount = it.get('amount', 0)
                try:
                    amount_f = float(amount)
                except Exception:
                    amount_f = 0.0

                combined[concept] = amount_f

            for concept, amount_f in combined.items():
                if not concept:
                    continue
                try:
                    amount_f = float(amount_f)
                except Exception:
                    amount_f = 0.0
                if amount_f == 0:
                    continue

                cursor.execute(
                    insert_query,
                    (
                        int(employee_id),
                        source_year_i,
                        source_month_i,
                        apply_year,
                        apply_month,
                        concept,
                        amount_f,
                    ),
                )

            connection.commit()
            return True
        except Exception as e:
            self.logger.error(f"Fehler beim Erstellen Carry Over: {e}")
            try:
                if connection is not None:
                    connection.rollback()
            except Exception:
                pass
            return False
        finally:
            if cursor is not None:
                try:
                    cursor.close()
                except Exception:
                    pass
            if connection is not None:
                try:
                    connection.close()
                except Exception:
                    pass

    def list_carry_over_by_source(
        self,
        employee_id: int,
        source_year: int,
        source_month: int,
    ) -> List[Dict]:
        try:
            query = """
            SELECT
                id_carry_over,
                id_empleado,
                source_anio,
                source_mes,
                apply_anio,
                apply_mes,
                concept,
                amount,
                created_at,
                updated_at
            FROM t010_carry_over
            WHERE id_empleado = %s
              AND source_anio = %s
              AND source_mes = %s
            ORDER BY created_at DESC, id_carry_over DESC
            """
            return self.execute_query(query, (int(employee_id), int(source_year), int(source_month)))
        except Exception as e:
            self.logger.error(f"Fehler beim Abrufen Carry Over (source): {e}")
            return []

    def delete_carry_over(self, carry_over_id: int) -> bool:
        try:
            query = "DELETE FROM t010_carry_over WHERE id_carry_over = %s"
            return self.execute_update(query, (int(carry_over_id),))
        except Exception as e:
            self.logger.error(f"Fehler beim Löschen Carry Over {carry_over_id}: {e}")
            return False

    def get_carry_over_sums_for_apply(
        self,
        apply_year: int,
        apply_month: int,
        employee_ids: List[int],
    ) -> List[Dict]:
        try:
            if not employee_ids:
                return []
            employee_ids_clean = [int(x) for x in employee_ids]
            placeholders = ", ".join(["%s"] * len(employee_ids_clean))
            query = f"""
            SELECT
                id_empleado,
                concept,
                SUM(amount) AS amount
            FROM t010_carry_over
            WHERE apply_anio = %s
              AND apply_mes = %s
              AND id_empleado IN ({placeholders})
            GROUP BY id_empleado, concept
            """
            params = (int(apply_year), int(apply_month), *employee_ids_clean)
            return self.execute_query(query, params)
        except Exception as e:
            self.logger.error(f"Fehler beim Aggregieren Carry Over: {e}")
            return []

    def disconnect(self):
        if self.connection and self.connection.is_connected():
            self.connection.close()
            self.logger.info("Datenbankverbindung geschlossen")

    def execute_query(self, query: str, params: tuple = None) -> List[Dict]:
        last_error: Optional[Error] = None
        for attempt in range(2):
            connection = None
            cursor = None
            try:
                connection = self._create_connection()
                cursor = connection.cursor(dictionary=True)
                cursor.execute(query, params)
                result = cursor.fetchall() if cursor.with_rows else []
                return result
            except Error as e:
                last_error = e
                if attempt == 0 and self._is_transient_connection_error(e):
                    self.logger.warning(f"Datenbankverbindung verloren, Retry... ({e})")
                    continue
                self.logger.error(f"Fehler bei der Abfrage: {e}")
                return []
            finally:
                if cursor is not None:
                    try:
                        cursor.close()
                    except Exception:
                        pass
                if connection is not None:
                    try:
                        connection.close()
                    except Exception:
                        pass
        if last_error is not None:
            self.logger.error(f"Fehler bei der Abfrage: {last_error}")
        return []

    def execute_update(self, query: str, params: tuple = None) -> bool:
        last_error: Optional[Error] = None
        for attempt in range(2):
            connection = None
            cursor = None
            try:
                connection = self._create_connection()
                cursor = connection.cursor()
                cursor.execute(query, params)
                connection.commit()
                return True
            except Error as e:
                last_error = e
                if attempt == 0 and self._is_transient_connection_error(e):
                    self.logger.warning(f"Datenbankverbindung verloren, Retry... ({e})")
                    continue
                self.logger.error(f"Fehler beim Update: {e}")
                self.logger.error(f"Query: {query}")
                self.logger.error(f"Params: {params}")
                try:
                    if connection is not None:
                        connection.rollback()
                except Exception:
                    pass
                return False
            finally:
                if cursor is not None:
                    try:
                        cursor.close()
                    except Exception:
                        pass
                if connection is not None:
                    try:
                        connection.close()
                    except Exception:
                        pass
        if last_error is not None:
            self.logger.error(f"Fehler beim Update: {last_error}")
        return False

    def get_payout_month(self) -> int:
        """Returns the globally configured payout month (1-12). Defaults to 4 (April)."""
        try:
            default_value = 4
            settings_path = os.path.join(os.path.dirname(__file__), "config", "settings.json")
            if not os.path.exists(settings_path):
                try:
                    with open(settings_path, "w", encoding="utf-8") as f:
                        json.dump({"payout_month": default_value}, f, ensure_ascii=False, indent=2)
                except Exception:
                    return default_value
                return default_value
            with open(settings_path, "r", encoding="utf-8") as f:
                data = json.load(f) or {}
            value = int(data.get("payout_month", default_value))
            if 1 <= value <= 12:
                return value
            return default_value
        except Exception as e:
            self.logger.error(f"Fehler beim Lesen von payout_month: {e}")
            return 4

    def set_payout_month(self, payout_month: int) -> bool:
        """Sets the globally configured payout month (1-12)."""
        try:
            if not isinstance(payout_month, int) or not (1 <= payout_month <= 12):
                return False
            settings_path = os.path.join(os.path.dirname(__file__), "config", "settings.json")
            with open(settings_path, "w", encoding="utf-8") as f:
                json.dump({"payout_month": payout_month}, f, ensure_ascii=False, indent=2)
            return True
        except Exception as e:
            self.logger.error(f"Fehler beim Setzen von payout_month: {e}")
            return False

    def get_all_employees(self) -> List[Dict]:
        query = """
        SELECT id_empleado, nombre, apellido, ceco, categoria, activo, fecha_alta, declaracion, dni
        FROM t001_empleados 
        ORDER BY apellido, nombre
        """
        return self.execute_query(query)

    def get_all_employees_with_salaries(self) -> List[Dict]:
        """Hole alle Mitarbeiter mit ihren Gehaltsdaten"""
        query = """
        SELECT e.id_empleado, e.nombre, e.apellido, e.ceco, e.categoria, e.activo, e.fecha_alta,
               s.anio, s.salario_anual_bruto, s.salario_mensual_bruto, 
               s.modalidad, s.atrasos, s.antiguedad
        FROM t001_empleados e
        LEFT JOIN t002_salarios s ON e.id_empleado = s.id_empleado
        ORDER BY e.apellido, e.nombre, s.anio DESC
        """
        employees_data = self.execute_query(query)
        # Gruppiere Gehaltsdaten pro Mitarbeiter
        employees = {}
        for row in employees_data:
            emp_id = row['id_empleado']
            if emp_id not in employees:
                employees[emp_id] = {
                    'id_empleado': row['id_empleado'],
                    'nombre': row['nombre'],
                    'apellido': row['apellido'],
                    'ceco': row['ceco'],
                    'categoria': row['categoria'],
                    'activo': row['activo'],
                    'fecha_alta': row.get('fecha_alta'),
                    'salaries': []
                }
            # Füge Gehaltsdaten hinzu, wenn vorhanden
            if row['anio'] is not None:
                employees[emp_id]['salaries'].append({
                    'anio': row['anio'],
                    'salario_anual_bruto': row['salario_anual_bruto'],
                    'salario_mensual_bruto': row['salario_mensual_bruto'],
                    'modalidad': row['modalidad'],
                    'atrasos': row['atrasos'],
                    'antiguedad': row['antiguedad']
                })
        return list(employees.values())

    def get_employee(self, employee_id: int) -> Optional[Dict]:
        """Hole einen Mitarbeiter anhand seiner ID"""
        try:
            query = """
            SELECT id_empleado, nombre, apellido, ceco, categoria, activo, fecha_alta, declaracion, dni
            FROM t001_empleados 
            WHERE id_empleado = %s
            """
            employees = self.execute_query(query, (employee_id,))
            return employees[0] if employees else None
        except Exception as e:
            self.logger.error(f"Fehler beim Abrufen des Mitarbeiters {employee_id}: {e}")
            return None

    def get_employee_complete_info(self, employee_id: int) -> Dict:
        # Mitarbeiterstammdaten
        employee_query = """
        SELECT id_empleado, nombre, apellido, ceco, categoria, activo, fecha_alta, declaracion, dni
        FROM t001_empleados 
        WHERE id_empleado = %s
        """
        employees = self.execute_query(employee_query, (employee_id,))
        if not employees:
            return {}
        employee = employees[0]
        # Gehaltsdaten
        salary_query = """
        SELECT anio, modalidad, antiguedad, salario_anual_bruto, salario_mensual_bruto, atrasos, salario_mensual_con_atrasos, fecha_modificacion
        FROM t002_salarios 
        WHERE id_empleado = %s 
        ORDER BY anio DESC
        """
        salaries = self.execute_query(salary_query, (employee_id,))
        # Bruttoeinkünfte (jahresabhängig) - Aggregiere aus monatlichen Daten
        ingresos_query = """
        SELECT 
            anio,
            ROUND(AVG(ticket_restaurant), 2) as ticket_restaurant,
            ROUND(AVG(primas), 2) as primas,
            ROUND(AVG(dietas_cotizables), 2) as dietas_cotizables,
            ROUND(AVG(horas_extras), 2) as horas_extras,
            ROUND(AVG(dias_exentos), 2) as dias_exentos,
            ROUND(AVG(dietas_exentas), 2) as dietas_exentas,
            ROUND(AVG(seguro_pensiones), 2) as seguro_pensiones,
            ROUND(AVG(lavado_coche), 2) as lavado_coche,
            ROUND(AVG(beca_escolar), 2) as beca_escolar,
            ROUND(AVG(formacion), 2) as formacion,
            ROUND(AVG(tickets), 2) as tickets,
            MAX(fecha_modificacion) as fecha_modificacion
        FROM t003_ingresos_brutos_mensuales 
        WHERE id_empleado = %s
        GROUP BY anio
        ORDER BY anio DESC
        """
        ingresos_results = self.execute_query(ingresos_query, (employee_id,))
        # Abzüge (jahresabhängig) - Aggregiere aus monatlichen Daten
        deducciones_query = """
        SELECT 
            anio,
            ROUND(AVG(seguro_accidentes), 2) as seguro_accidentes,
            ROUND(AVG(adelas), 2) as adelas,
            ROUND(AVG(sanitas), 2) as sanitas,
            ROUND(AVG(gasolina), 2) as gasolina,
            ROUND(AVG(ret_especie), 2) as ret_especie,
            ROUND(AVG(seguro_medico), 2) as seguro_medico,
            ROUND(AVG(cotizacion_especie), 2) as cotizacion_especie,
            MAX(fecha_modificacion) as fecha_modificacion
        FROM t004_deducciones_mensuales 
        WHERE id_empleado = %s
        GROUP BY anio
        ORDER BY anio DESC
        """
        deducciones_results = self.execute_query(deducciones_query, (employee_id,))
        # Monatliche Bruttoeinkünfte
        ingresos_mensuales_query = """
        SELECT anio, mes, ticket_restaurant, primas, 
               dietas_cotizables, horas_extras, dias_exentos, 
               dietas_exentas, seguro_pensiones, lavado_coche, beca_escolar, formacion, tickets, fecha_modificacion
        FROM t003_ingresos_brutos_mensuales 
        WHERE id_empleado = %s
        ORDER BY anio DESC, mes ASC
        """
        ingresos_mensuales_results = self.execute_query(ingresos_mensuales_query, (employee_id,))
        # Monatliche Abzüge
        deducciones_mensuales_query = """
        SELECT anio, mes, seguro_accidentes, adelas, sanitas, 
               gasolina, ret_especie, seguro_medico, cotizacion_especie, fecha_modificacion
        FROM t004_deducciones_mensuales 
        WHERE id_empleado = %s
        ORDER BY anio DESC, mes ASC
        """
        deducciones_mensuales_results = self.execute_query(deducciones_mensuales_query, (employee_id,))
        fte_query = """
        SELECT anio, mes, porcentaje, fecha_modificacion
        FROM t008_empleado_fte
        WHERE id_empleado = %s
        ORDER BY anio DESC, mes DESC
        """
        fte_results = self.execute_query(fte_query, (employee_id,))
        return {
            'employee': employee,
            'salaries': salaries,
            'ingresos': ingresos_results,
            'deducciones': deducciones_results,
            'ingresos_mensuales': ingresos_mensuales_results,
            'deducciones_mensuales': deducciones_mensuales_results,
            'fte': fte_results,
        }

    def get_employee_fte(self, employee_id: int) -> List[Dict[str, Any]]:
        try:
            query = """
            SELECT anio, mes, porcentaje, fecha_modificacion
            FROM t008_empleado_fte
            WHERE id_empleado = %s
            ORDER BY anio DESC, mes DESC
            """
            return self.execute_query(query, (employee_id,))
        except Exception as e:
            self.logger.error(f"Fehler beim Abrufen FTE für Mitarbeiter {employee_id}: {e}")
            return []

    def upsert_employee_fte(self, employee_id: int, year: int, month: int, porcentaje: float) -> bool:
        try:
            year = int(year)
            month = int(month)
            porcentaje = float(porcentaje)
            if year < 2000 or year > 2100:
                return False
            if month < 1 or month > 12:
                return False
            if porcentaje < 0 or porcentaje > 100:
                return False
            query = """
            INSERT INTO t008_empleado_fte (id_empleado, anio, mes, porcentaje)
            VALUES (%s, %s, %s, %s)
            ON DUPLICATE KEY UPDATE
                porcentaje = VALUES(porcentaje)
            """
            return self.execute_update(query, (employee_id, year, month, porcentaje))
        except Exception as e:
            self.logger.error(f"Fehler beim Upsert FTE für Mitarbeiter {employee_id}: {e}")
            return False

    def delete_employee_fte(self, employee_id: int, year: int, month: int) -> bool:
        try:
            year = int(year)
            month = int(month)
            query = """
            DELETE FROM t008_empleado_fte
            WHERE id_empleado = %s AND anio = %s AND mes = %s
            """
            return self.execute_update(query, (employee_id, year, month))
        except Exception as e:
            self.logger.error(f"Fehler beim Löschen FTE für Mitarbeiter {employee_id}: {e}")
            return False

    def get_employee_fte_effective_percent(self, employee_id: int, year: int, month: int) -> float:
        try:
            year = int(year)
            month = int(month)
            query = """
            SELECT porcentaje
            FROM t008_empleado_fte
            WHERE id_empleado = %s
              AND (anio < %s OR (anio = %s AND mes <= %s))
            ORDER BY anio DESC, mes DESC
            LIMIT 1
            """
            rows = self.execute_query(query, (employee_id, year, year, month))
            if rows and rows[0].get('porcentaje') is not None:
                return float(rows[0]['porcentaje'])
            return 100.0
        except Exception as e:
            self.logger.error(f"Fehler beim Lesen effective FTE für Mitarbeiter {employee_id}: {e}")
            return 100.0

    def update_employee(self, employee_id: int, table: str, data: Dict[str, Any]) -> bool:
        if table == 't001_empleados':
            # Nur updatable Felder erlauben
            allowed_fields = ['nombre', 'apellido', 'ceco', 'categoria', 'activo', 'fecha_alta', 'declaracion', 'dni']
            update_fields = []
            update_values = []
            for field, value in data.items():
                if field in allowed_fields:
                    if field == 'categoria' and not self._is_valid_employee_category(value):
                        continue
                    if field == 'categoria':
                        value = self._normalize_employee_category(value)
                    if field == 'declaracion' and value is not None:
                        value = str(value).strip().upper()
                    update_fields.append(f"{field} = %s")
                    update_values.append(value)
            if not update_fields:
                return False
            query = f"""
            UPDATE t001_empleados 
            SET {', '.join(update_fields)} 
            WHERE id_empleado = %s
            """
            update_values.append(employee_id)
            return self.execute_update(query, tuple(update_values))
        return False

    def update_salary(self, employee_id: int, year: int, data: Dict[str, Any]) -> bool:
        # Prüfen ob Gehalt für dieses Jahr existiert
        check_query = """
        SELECT id_empleado FROM t002_salarios 
        WHERE id_empleado = %s AND anio = %s
        """
        exists = self.execute_query(check_query, (employee_id, year))
        allowed_fields = ['modalidad', 'salario_anual_bruto', 'antiguedad']
        if exists:
            # Update
            update_fields = []
            update_values = []
            for field, value in data.items():
                if field in allowed_fields:
                    update_fields.append(f"{field} = %s")
                    update_values.append(value)
            if not update_fields:
                return False
            # Wenn sich das Jahresgehalt ändert, auch das Monatsgehalt neu berechnen
            if 'salario_anual_bruto' in data:
                current_salary = float(data['salario_anual_bruto'])
                modalidad = int(data.get('modalidad', 12))
                # Berechne atrasos mit neuer Python-Funktion
                calculated_atrasos = self.calculate_atrasos(employee_id, year, modalidad, current_salary)
                # Füge atrasos und salario_mensual_con_atrasos zum Update hinzu
                update_fields.extend(["atrasos = %s", "salario_mensual_con_atrasos = %s"])
                # Wenn keine Atrasos (kein Vorjahresgehalt), dann salario_mensual_con_atrasos = salario_mensual_bruto
                if calculated_atrasos > 0:
                    update_values.extend([calculated_atrasos, (current_salary / modalidad) + calculated_atrasos])
                else:
                    update_values.extend([0.0, current_salary / modalidad])
                # Berechne salario_mensual_bruto
                update_fields.append("salario_mensual_bruto = %s")
                update_values.append(current_salary / modalidad)
            query = f"""
            UPDATE t002_salarios 
            SET {', '.join(update_fields)} 
            WHERE id_empleado = %s AND anio = %s
            """
            update_values.extend([employee_id, year])
            result = self.execute_update(query, tuple(update_values))
            # Wenn erfolgreich und sich das Jahresgehalt geändert hat, aktualisiere Folgejahre
            if result and 'salario_anual_bruto' in data:
                self._update_subsequent_years_atrasos(employee_id, year)
            return result
        else:
            # Insert
            insert_fields = ['id_empleado', 'anio'] + [field for field in data.keys() if field in allowed_fields]
            insert_values = [employee_id, year] + [data[field] for field in insert_fields[2:]]
            # Berechne Monatsgehalt und atrasos wenn Jahresgehalt angegeben ist
            if 'salario_anual_bruto' in data:
                current_salary = float(data['salario_anual_bruto'])
                modalidad = int(data.get('modalidad', 12))
                # Berechne atrasos mit neuer Python-Funktion
                calculated_atrasos = self.calculate_atrasos(employee_id, year, modalidad, current_salary)
                # Füge berechnete Felder hinzu
                insert_fields.extend(['salario_mensual_bruto', 'atrasos', 'salario_mensual_con_atrasos'])
                monthly_salary = current_salary / modalidad
                # Wenn keine Atrasos (kein Vorjahresgehalt), dann salario_mensual_con_atrasos = salario_mensual_bruto
                if calculated_atrasos > 0:
                    insert_values.extend([monthly_salary, calculated_atrasos, monthly_salary + calculated_atrasos])
                else:
                    insert_values.extend([monthly_salary, 0.0, monthly_salary])
            placeholders = ', '.join(['%s'] * len(insert_fields))
            query = f"""
            INSERT INTO t002_salarios ({', '.join(insert_fields)}) 
            VALUES ({placeholders})
            """
            result = self.execute_update(query, tuple(insert_values))
            # Wenn erfolgreich, aktualisiere Folgejahre
            if result:
                self._update_subsequent_years_atrasos(employee_id, year)
            return result

    def search_employees(self, search_term: str) -> List[Dict]:
        query = """
        SELECT id_empleado, nombre, apellido, ceco, activo 
        FROM t001_empleados 
        WHERE id_empleado = %s 
           OR LOWER(nombre) LIKE LOWER(%s) 
           OR LOWER(apellido) LIKE LOWER(%s) 
           OR LOWER(ceco) LIKE LOWER(%s)
        ORDER BY apellido, nombre
        """
        search_pattern = f"%{search_term}%"
        return self.execute_query(query, (search_term, search_pattern, search_pattern, search_pattern))

    def add_employee(self, employee_data: Dict[str, Any]) -> int:
        try:
            # Neue ID holen
            max_id_query = "SELECT COALESCE(MAX(id_empleado), 0) as max_id FROM t001_empleados"
            result = self.execute_query(max_id_query)
            new_id = result[0]['max_id'] + 1 if result else 1
            # Mitarbeiter einfügen
            if not self._is_valid_employee_category(employee_data.get('categoria')):
                return -1
            insert_query = """
            INSERT INTO t001_empleados (id_empleado, nombre, apellido, ceco, categoria, activo, fecha_alta, declaracion, dni)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
            """
            declaracion = employee_data.get('declaracion')
            declaracion = str(declaracion).strip().upper() if declaracion is not None and str(declaracion).strip() != '' else None
            params = (
                new_id,
                employee_data.get('nombre', ''),
                employee_data.get('apellido', ''),
                employee_data.get('ceco', ''),
                self._normalize_employee_category(employee_data.get('categoria')),
                employee_data.get('activo', True),
                employee_data.get('fecha_alta'),
                declaracion,
                employee_data.get('dni'),
            )
            if self.execute_update(insert_query, params):
                # Standard-Datensätze für neuen Mitarbeiter erstellen
                self._create_default_records(new_id)
                return new_id
            else:
                return -1
        except Exception as e:
            self.logger.error(f"Fehler beim Hinzufügen des Mitarbeiters: {e}")
            return -1

    def _create_default_records(self, employee_id: int):
        # Nur noch monatliche Standarddatensätze für das aktuelle Jahr erstellen
        current_year = datetime.now().year
        self._create_monthly_default_records(employee_id, current_year)

    def _create_monthly_default_records(self, employee_id: int, year: int):
        """Erstellt Standard-Monatsdatensätze für einen Mitarbeiter"""
        try:
            # Monatliche Bruttoeinkünfte für alle 12 Monate erstellen
            for month in range(1, 13):
                ingresos_monthly_query = """
                INSERT INTO t003_ingresos_brutos_mensuales (id_empleado, anio, mes, ticket_restaurant, 
                                                           primas, dietas_cotizables, horas_extras, dias_exentos, 
                                                           dietas_exentas, seguro_pensiones, lavado_coche, formacion, tickets) 
                VALUES (%s, %s, %s, 0, 0, 0, 0, 0, 0, 0, 0, 0, 0)
                ON DUPLICATE KEY UPDATE
                    ticket_restaurant = VALUES(ticket_restaurant),
                    primas = VALUES(primas),
                    dietas_cotizables = VALUES(dietas_cotizables),
                    horas_extras = VALUES(horas_extras),
                    dias_exentos = VALUES(dias_exentos),
                    dietas_exentas = VALUES(dietas_exentas),
                    seguro_pensiones = VALUES(seguro_pensiones),
                    lavado_coche = VALUES(lavado_coche),
                    formacion = VALUES(formacion),
                    tickets = VALUES(tickets)
                """
                self.execute_update(ingresos_monthly_query, (employee_id, year, month))
            # Monatliche Abzüge für alle 12 Monate erstellen
            for month in range(1, 13):
                deducciones_monthly_query = """
                INSERT INTO t004_deducciones_mensuales (id_empleado, anio, mes, seguro_accidentes, 
                                                        adelas, sanitas, gasolina, cotizacion_especie) 
                VALUES (%s, %s, %s, 0, 0, 0, 0, 0)
                ON DUPLICATE KEY UPDATE
                    seguro_accidentes = VALUES(seguro_accidentes),
                    adelas = VALUES(adelas),
                    sanitas = VALUES(sanitas),
                    gasolina = VALUES(gasolina),
                    cotizacion_especie = VALUES(cotizacion_especie)
                """
                self.execute_update(deducciones_monthly_query, (employee_id, year, month))
        except Exception as e:
            self.logger.error(f"Fehler beim Erstellen der monatlichen Standarddatensätze: {e}")
    def add_salary(self, employee_id: int, salary_data: Dict[str, Any]) -> bool:
        try:
            # Prüfen ob Mitarbeiter existiert
            employee_query = "SELECT id_empleado FROM t001_empleados WHERE id_empleado = %s"
            employee = self.execute_query(employee_query, (employee_id,))
            if not employee:
                self.logger.error(f"Mitarbeiter mit ID {employee_id} nicht gefunden")
                return False
            # Prüfen ob Gehalt für dieses Jahr bereits existiert
            year = salary_data.get('anio')
            if year is None:
                self.logger.error("Jahr nicht angegeben")
                return False
            check_query = "SELECT id_empleado FROM t002_salarios WHERE id_empleado = %s AND anio = %s"
            exists = self.execute_query(check_query, (employee_id, year))
            if exists:
                self.logger.error(f"Gehalt für Jahr {year} existiert bereits für Mitarbeiter {employee_id}")
                return False
            # Neues Gehalt einfügen
            # Berechne atrasos mit neuer Python-Funktion
            calculated_atrasos = self.calculate_atrasos(employee_id, year, salary_data.get('modalidad', 12), salary_data.get('salario_anual_bruto', 0))
            insert_query = """
            INSERT INTO t002_salarios (id_empleado, anio, modalidad, antiguedad, 
                                     salario_anual_bruto, salario_mensual_bruto, 
                                     atrasos, salario_mensual_con_atrasos) 
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
            """
            salario_anual = salary_data.get('salario_anual_bruto', 0)
            modalidad = salary_data.get('modalidad', 12)
            salario_mensual = salario_anual / modalidad
            # Wenn keine Atrasos (kein Vorjahresgehalt), dann salario_mensual_con_atrasos = salario_mensual
            if calculated_atrasos > 0:
                salario_con_atrasos = salario_mensual + calculated_atrasos
            else:
                salario_con_atrasos = salario_mensual
            params = (
                employee_id,
                year,
                salary_data.get('modalidad', 12),
                salary_data.get('antiguedad', 0),
                salario_anual,
                salario_mensual,
                calculated_atrasos,
                salario_con_atrasos
            )
            result = self.execute_update(insert_query, params)
            # Wenn erfolgreich, prüfe ob Folgejahre existieren und aktualisiere ihre atrasos
            if result:
                self._update_subsequent_years_atrasos(employee_id, year)
            return result
        except Exception as e:
            self.logger.error(f"Fehler beim Hinzufügen des Gehalts: {e}")
            return False
    def get_salary(self, employee_id: int, year: int) -> Dict[str, Any] | None:
        """Holt Gehaltsdaten für einen Mitarbeiter und Jahr"""
        try:
            query = """
            SELECT id_empleado, anio, salario_anual_bruto, salario_mensual_bruto, 
                   modalidad, antiguedad, atrasos, salario_mensual_con_atrasos, 
                   fecha_creacion, fecha_modificacion
            FROM t002_salarios 
            WHERE id_empleado = %s AND anio = %s
            """
            rows = self.execute_query(query, (employee_id, year))
            if rows and len(rows) > 0:
                return rows[0]
            return None
        except Exception as e:
            self.logger.error(f"Fehler beim Abrufen des Gehalts für Mitarbeiter {employee_id}, Jahr {year}: {e}")
            return None
    def apply_employee_salary_increase(self, employee_id: int, target_year: int, percentage_increase: float) -> Dict[str, Any]:
        """
        Wendet eine prozentuale Gehaltserhöhung auf einen einzelnen Mitarbeiter an.
        Die Erhöhung wird erst im April des target_year wirksam.
        Args:
            employee_id: ID des Mitarbeiters
            target_year: Jahr in dem die Erhöhung wirksam wird
            percentage_increase: Prozentsatz der Erhöhung (z.B. 10.0 für 10%)
        Returns:
            Dict mit Ergebnissen der Operation
        """
        try:
            # Hole Mitarbeiterdaten für den Log-Eintrag
            employee_info = self.get_employee(employee_id)
            if not employee_info:
                return {"success": False, "message": f"Mitarbeiter {employee_id} nicht gefunden"}
            
            # Zuerst prüfen ob bereits ein Gehalt für target_year existiert
            existing_salary_query = """
            SELECT salario_anual_bruto, modalidad
            FROM t002_salarios 
            WHERE id_empleado = %s AND anio = %s
            """
            existing_salary = self.execute_query(existing_salary_query, (employee_id, target_year))
            if existing_salary and existing_salary[0]['salario_anual_bruto'] is not None and float(existing_salary[0]['salario_anual_bruto']) > 0:
                # Gehalt für target_year existiert und ist > 0 - dieses als Grundlage nehmen
                current_salary = float(existing_salary[0]['salario_anual_bruto'])
                modalidad = int(existing_salary[0]['modalidad'])
                base_year = target_year
            else:
                # Kein gültiges Gehalt für target_year (NULL oder 0) - nehme Gehalt aus dem höchsten verfügbaren Jahr mit Gehalt > 0
                latest_salary_query = """
                SELECT e.id_empleado, e.nombre, e.apellido, 
                       COALESCE(s.salario_anual_bruto, 0) as salario_actual,
                       COALESCE(s.modalidad, 12) as modalidad,
                       COALESCE(s.anio, %s) as salary_year
                FROM t001_empleados e
                LEFT JOIN t002_salarios s ON e.id_empleado = s.id_empleado 
                WHERE e.id_empleado = %s 
                ORDER BY s.anio DESC
                LIMIT 1
                """
                employees = self.execute_query(latest_salary_query, (target_year - 1, employee_id))
                if not employees:
                    return {"success": False, "message": f"Kein Gehalt für Mitarbeiter {employee_id} gefunden"}
                employee = employees[0]
                current_salary = float(employee['salario_actual'])
                modalidad = int(employee['modalidad'])
                base_year = employee.get('salary_year', target_year - 1)
                # Wenn das gefundene Gehalt auch 0 ist, suche nach dem letzten Gehalt > 0
                if current_salary == 0:
                    fallback_salary_query = """
                    SELECT salario_anual_bruto, modalidad, anio
                    FROM t002_salarios 
                    WHERE id_empleado = %s AND salario_anual_bruto > 0
                    ORDER BY anio DESC
                    LIMIT 1
                    """
                    fallback_salary = self.execute_query(fallback_salary_query, (employee_id,))
                    if fallback_salary:
                        current_salary = float(fallback_salary[0]['salario_anual_bruto'])
                        modalidad = int(fallback_salary[0]['modalidad'])
                        base_year = fallback_salary[0]['anio']
            # Berechne neues Gehalt
            new_salary = current_salary * (1 + percentage_increase / 100)
            # Berechne atrasos korrekt für April-Wirksamkeit
            # Nachzahlung für die Monate Januar-März mit der Differenz zwischen alt und neu
            old_monthly = current_salary / modalidad
            new_monthly = new_salary / modalidad
            monthly_diff = new_monthly - old_monthly
            payout_month = self.get_payout_month()
            months_before_payout = max(0, payout_month - 1)
            atrasos = monthly_diff * months_before_payout
            # Füge oder aktualisiere Gehalt für target_year
            salary_data = {
                'anio': target_year,  # Jahr muss im salary_data enthalten sein
                'salario_anual_bruto': new_salary,
                'modalidad': modalidad,
                'antiguedad': 0
            }
            # Versuche zuerst zu aktualisieren, wenn nicht vorhanden dann hinzufügen
            result = self.update_salary(employee_id, target_year, salary_data)
            if not result:
                # Wenn kein Gehalt für dieses Jahr existiert, füge es hinzu
                result = self.add_salary(employee_id, salary_data)
            if result:
                # Manuelles Update der atrasos mit korrekter Berechnung
                update_atrasos_query = """
                UPDATE t002_salarios 
                SET atrasos = %s, 
                    salario_mensual_con_atrasos = %s
                WHERE id_empleado = %s AND anio = %s
                """
                new_monthly_salary = new_salary / modalidad
                # Atrasos werden nur gespeichert, aber nicht zum regulären Monatsgehalt addiert
                # Sie werden nur einmalig im April gezahlt
                self.execute_update(update_atrasos_query, (
                    atrasos,
                    new_monthly_salary,  # Normales Monatsgehalt ohne atrasos
                    employee_id,
                    target_year
                ))
                return {
                    "success": True,
                    "updated_count": 1,
                    "error_count": 0,
                    "employees": [{
                        'id': employee_id,
                        'name': f"{employee_info['nombre']} {employee_info['apellido']}",
                        'old_salary': current_salary,
                        'new_salary': new_salary,
                        'increase_percent': percentage_increase,
                        'atrasos': atrasos,
                        'base_year': base_year
                    }],
                    "errors": [],
                    "message": f"Gehaltserhöhung für {employee_info['nombre']} {employee_info['apellido']} erfolgreich durchgeführt (Basis: {base_year})"
                }
            else:
                return {"success": False, "message": f"Fehler bei Aktualisierung von {employee_info['nombre']} {employee_info['apellido']}"}
        except Exception as e:
            self.logger.error(f"Fehler bei prozentualer Gehaltserhöhung für Mitarbeiter {employee_id}: {e}")
            return {"success": False, "message": f"Datenbankfehler: {str(e)}"}
    def apply_percentage_salary_increase(self, target_year: int, percentage_increase: float = None, absolute_increase: float = None, excluded_employee_ids: List[int] = None) -> Dict[str, Any]:
        """
        Wendet eine Gehaltserhöhung auf alle aktiven Mitarbeiter an.
        Die Erhöhung wird erst im April des target_year wirksam.
        Args:
            target_year: Jahr in dem die Erhöhung wirksam wird
            percentage_increase: Prozentsatz der Erhöhung (z.B. 10.0 für 10%)
            absolute_increase: Absoluter Betrag der Erhöhung (z.B. 1000.0 für 1000€)
            excluded_employee_ids: Liste von Mitarbeiter-IDs, die von der Erhöhung ausgeschlossen werden sollen
        Returns:
            Dict mit Ergebnissen der Operation
        """
        try:
            if excluded_employee_ids is None:
                excluded_employee_ids = []
            if percentage_increase is None and absolute_increase is None:
                return {"success": False, "message": "Entweder percentage_increase oder absolute_increase muss angegeben werden"}
            # Hole alle aktiven Mitarbeiter
            query = """
            SELECT e.id_empleado, e.nombre, e.apellido 
            FROM t001_empleados e
            WHERE e.activo = TRUE
            """
            if excluded_employee_ids:
                # Schließe ausgeschlossene Mitarbeiter aus
                placeholders = ','.join(['%s'] * len(excluded_employee_ids))
                query += f" AND e.id_empleado NOT IN ({placeholders})"
                params = tuple(excluded_employee_ids)
            else:
                params = ()
            employees = self.execute_query(query, params)
            if not employees:
                return {"success": False, "message": "Keine Mitarbeiter gefunden"}
            updated_employees = []
            errors = []
            for employee in employees:
                try:
                    employee_id = employee['id_empleado']
                    # Für jeden Mitarbeiter prüfen ob bereits ein Gehalt für target_year existiert
                    existing_salary_query = """
                    SELECT salario_anual_bruto, modalidad
                    FROM t002_salarios 
                    WHERE id_empleado = %s AND anio = %s
                    """
                    existing_salary = self.execute_query(existing_salary_query, (employee_id, target_year))
                    if existing_salary and existing_salary[0]['salario_anual_bruto'] is not None and float(existing_salary[0]['salario_anual_bruto']) > 0:
                        # Gehalt für target_year existiert und ist > 0 - dieses als Grundlage nehmen
                        current_salary = float(existing_salary[0]['salario_anual_bruto'])
                        modalidad = int(existing_salary[0]['modalidad'])
                        base_year = target_year
                    else:
                        # Kein gültiges Gehalt für target_year (NULL oder 0) - nehme Gehalt aus dem höchsten verfügbaren Jahr mit Gehalt > 0
                        latest_salary_query = """
                        SELECT COALESCE(s.salario_anual_bruto, 0) as salario_actual,
                               COALESCE(s.modalidad, 12) as modalidad,
                               COALESCE(s.anio, %s) as salary_year
                        FROM t002_salarios s
                        WHERE s.id_empleado = %s 
                        ORDER BY s.anio DESC
                        LIMIT 1
                        """
                        salary_data = self.execute_query(latest_salary_query, (target_year - 1, employee_id))
                        if not salary_data:
                            errors.append(f"Mitarbeiter {employee['nombre']} {employee['apellido']}: Kein Gehalt gefunden")
                            continue
                        current_salary = float(salary_data[0]['salario_actual'])
                        modalidad = int(salary_data[0]['modalidad'])
                        base_year = salary_data[0]['salary_year']
                        # Wenn das gefundene Gehalt auch 0 ist, suche nach dem letzten Gehalt > 0
                        if current_salary == 0:
                            fallback_salary_query = """
                            SELECT salario_anual_bruto, modalidad, anio
                            FROM t002_salarios 
                            WHERE id_empleado = %s AND salario_anual_bruto > 0
                            ORDER BY anio DESC
                            LIMIT 1
                            """
                            fallback_salary = self.execute_query(fallback_salary_query, (employee_id,))
                            if fallback_salary:
                                current_salary = float(fallback_salary[0]['salario_anual_bruto'])
                                modalidad = int(fallback_salary[0]['modalidad'])
                                base_year = fallback_salary[0]['anio']
                    # Berechne neues Gehalt
                    if percentage_increase is not None:
                        # Prozentsuale Erhöhung
                        new_salary = current_salary * (1 + percentage_increase / 100)
                        increase_info = f"+{percentage_increase}%"
                    else:
                        # Absolute Betragserhöhung
                        new_salary = current_salary + absolute_increase
                        increase_info = f"+{absolute_increase}€"
                    # Berechne atrasos korrekt für April-Wirksamkeit
                    old_monthly = current_salary / modalidad
                    new_monthly = new_salary / modalidad
                    monthly_diff = new_monthly - old_monthly
                    payout_month = self.get_payout_month()
                    months_before_payout = max(0, payout_month - 1)
                    atrasos = monthly_diff * months_before_payout
                    # Füge oder aktualisiere Gehalt für target_year
                    salary_data = {
                        'anio': target_year,  # Jahr muss im salary_data enthalten sein
                        'salario_anual_bruto': new_salary,
                        'modalidad': modalidad,
                        'antiguedad': 0
                    }
                    # Versuche zuerst zu aktualisieren, wenn nicht vorhanden dann hinzufügen
                    result = self.update_salary(employee_id, target_year, salary_data)
                    if not result:
                        # Wenn kein Gehalt für dieses Jahr existiert, füge es hinzu
                        result = self.add_salary(employee_id, salary_data)
                    if result:
                        # Manuelles Update der atrasos mit korrekter Berechnung
                        update_atrasos_query = """
                        UPDATE t002_salarios 
                        SET atrasos = %s, 
                            salario_mensual_con_atrasos = %s
                        WHERE id_empleado = %s AND anio = %s
                        """
                        new_monthly_salary = new_salary / modalidad
                        # Atrasos werden nur gespeichert, aber nicht zum regulären Monatsgehalt addiert
                        # Sie werden nur einmalig im April gezahlt
                        self.execute_update(update_atrasos_query, (
                            atrasos,
                            new_monthly_salary,  # Normales Monatsgehalt ohne atrasos
                            employee_id,
                            target_year
                        ))
                        updated_employees.append({
                            'id': employee_id,
                            'name': f"{employee['nombre']} {employee['apellido']}",
                            'old_salary': current_salary,
                            'new_salary': new_salary,
                            'increase_percent': percentage_increase if percentage_increase is not None else None,
                            'increase_absolute': absolute_increase if absolute_increase is not None else None,
                            'increase_info': increase_info,
                            'atrasos': atrasos,
                            'base_year': base_year
                        })
                    else:
                        errors.append(f"Fehler bei Aktualisierung von {employee['nombre']} {employee['apellido']}")
                except Exception as e:
                    errors.append(f"Mitarbeiter {employee['nombre']} {employee['apellido']}: {str(e)}")
            return {
                "success": len(updated_employees) > 0,
                "updated_count": len(updated_employees),
                "error_count": len(errors),
                "employees": updated_employees,
                "errors": errors,
                "message": f"{len(updated_employees)} Mitarbeiter erfolgreich aktualisiert"
            }
        except Exception as e:
            self.logger.error(f"Fehler bei prozentualer Gehaltserhöhung: {e}")
            return {"success": False, "message": f"Datenbankfehler: {str(e)}"}
    def calculate_atrasos(self, employee_id: int, year: int, modalidad: int, current_salary: float) -> float:
        """
        Berechnet atrasos nach gleicher Logik wie die Datenbank-Trigger.
        Args:
            employee_id: Mitarbeiter-ID
            year: Jahr für das die atrasos berechnet werden
            modalidad: Anzahl Gehaltszahlungen (12 oder 14)
            current_salary: Aktuelles Jahresgehalt
        Returns:
            Berechnete atrasos (immer >= 0)
        """
        try:
            # Prüfe ob Gehalt für Vorjahr existiert
            prev_year_query = """
            SELECT salario_anual_bruto 
            FROM t002_salarios 
            WHERE id_empleado = %s AND anio = %s
            """
            prev_year_data = self.execute_query(prev_year_query, (employee_id, year - 1))
            if not prev_year_data:
                # Kein Vorjahresgehalt vorhanden (neuer Mitarbeiter)
                return 0.0
            prev_salary = float(prev_year_data[0]['salario_anual_bruto'])
            if prev_salary <= 0:
                # Vorjahresgehalt ist 0 oder weniger (wie kein Gehalt)
                return 0.0
            payout_month = self.get_payout_month()
            months_before_payout = max(0, payout_month - 1)
            # Berechne atrasos nach Trigger-Logik
            if modalidad == 12:
                atrasos = (current_salary - prev_salary) / 12 * months_before_payout
            elif modalidad == 14:
                atrasos = (current_salary - prev_salary) / 14 * months_before_payout
            else:
                # Standard für unbekannte modalidad
                atrasos = (current_salary - prev_salary) / 12 * months_before_payout
            # Atrasos dürfen nie negativ sein
            return max(0.0, atrasos)
        except Exception as e:
            self.logger.error(f"Fehler bei atrasos-Berechnung: {e}")
            return 0.0
    def _update_subsequent_years_atrasos(self, employee_id: int, base_year: int) -> None:
        """Aktualisiert atrasos für alle Folgejahre eines Mitarbeiters"""
        try:
            # Finde alle Jahre größer als base_year für diesen Mitarbeiter
            query = """
            SELECT anio, modalidad, salario_anual_bruto 
            FROM t002_salarios 
            WHERE id_empleado = %s AND anio > %s 
            ORDER BY anio ASC
            """
            subsequent_years = self.execute_query(query, (employee_id, base_year))
            # Finde das Gehalt des base_year, um die ursprüngliche Gehaltserhöhung zu identifizieren
            base_year_query = """
            SELECT salario_anual_bruto 
            FROM t002_salarios 
            WHERE id_empleado = %s AND anio = %s
            """
            base_year_data = self.execute_query(base_year_query, (employee_id, base_year))
            base_year_salary = base_year_data[0]['salario_anual_bruto'] if base_year_data else 0
            for year_data in subsequent_years:
                current_year = year_data['anio']
                modalidad = year_data['modalidad']
                current_salario = year_data['salario_anual_bruto']
                # Finde das Gehalt des Vorjahres
                prev_query = """
                SELECT salario_anual_bruto 
                FROM t002_salarios 
                WHERE id_empleado = %s AND anio = %s
                """
                prev_year_data = self.execute_query(prev_query, (employee_id, current_year - 1))
                if prev_year_data:
                    prev_salario = prev_year_data[0]['salario_anual_bruto']
                    # Berechne atrasos: Nur für das erste Jahr nach der Gehaltserhöhung, wenn sich das Gehalt wirklich geändert hat
                    # Wenn das aktuelle Gehalt gleich dem Gehalt des base_year ist, gibt es keine neue Gehaltserhöhung
                    # Wenn Vorjahresgehalt 0 ist, werden keine atrasos berechnet (wie kein Gehalt)
                    if current_year == base_year + 1 and current_salario != prev_salario and float(prev_salario) > 0:
                        # Erste Gehaltserhöhung - berechne atrasos für Januar-März
                        if modalidad == 12:
                            payout_month = self.get_payout_month()
                            months_before_payout = max(0, payout_month - 1)
                            new_atrasos = (current_salario - prev_salario) / 12 * months_before_payout
                        elif modalidad == 14:
                            payout_month = self.get_payout_month()
                            months_before_payout = max(0, payout_month - 1)
                            new_atrasos = (current_salario - prev_salario) / 14 * months_before_payout
                        else:
                            new_atrasos = 0
                    else:
                        # Folgejahre oder keine Gehaltsänderung - keine atrasos
                        new_atrasos = 0
                    # Berechne neues salario_mensual_bruto
                    new_salario_mensual = current_salario / modalidad if modalidad in [12, 14] else current_salario / 12
                    # Aktualisiere den Datensatz
                    update_query = """
                    UPDATE t002_salarios 
                    SET atrasos = %s, 
                        salario_mensual_con_atrasos = %s,
                        salario_mensual_bruto = %s
                    WHERE id_empleado = %s AND anio = %s
                    """
                    self.execute_update(update_query, (
                        new_atrasos,
                        new_salario_mensual,  # Normales Monatsgehalt ohne atrasos
                        new_salario_mensual,
                        employee_id,
                        current_year
                    ))
        except Exception as e:
            self.logger.error(f"Fehler bei der Aktualisierung der Folgejahre: {e}")
    def update_ingresos(self, employee_id: int, year: int, data: Dict[str, Any]) -> bool:
        """Aktualisiert oder fügt Bruttoeinkünfte für einen Mitarbeiter für alle Monate eines Jahres hinzu"""
        try:
            self.logger.info(f"update_ingresos aufgerufen für employee_id={employee_id}, year={year}")
            allowed_fields = [
                'ticket_restaurant', 'primas', 
                'dietas_cotizables', 'horas_extras', 'dias_exentos', 
                'dietas_exentas', 'seguro_pensiones', 'lavado_coche', 'formacion', 'tickets'
                ]
            # beca_escolar wird im Jahresmodus nicht unterstützt; wird aus Daten entfernt
            filtered_data = {k: v for k, v in data.items() if k in allowed_fields}
            if not filtered_data:
                return False
            # Für jeden Monat (1-12) die monatlichen Datensätze aktualisieren
            success_all = True
            for month in range(1, 13):
                # Prüfen ob monatlicher Datensatz existiert
                check_query = """
                SELECT id_empleado FROM t003_ingresos_brutos_mensuales 
                WHERE id_empleado = %s AND anio = %s AND mes = %s
                """
                exists = self.execute_query(check_query, (employee_id, year, month))
                if exists:
                    # Update monatlicher Datensatz
                    update_fields = []
                    update_values = []
                    for field, value in filtered_data.items():
                        update_fields.append(f"{field} = %s")
                        update_values.append(value)
                    query = f"""
                    UPDATE t003_ingresos_brutos_mensuales 
                    SET {', '.join(update_fields)} 
                    WHERE id_empleado = %s AND anio = %s AND mes = %s
                    """
                    update_values.extend([employee_id, year, month])
                    result = self.execute_update(query, tuple(update_values))
                    if not result:
                        success_all = False
                        self.logger.error(f"Fehler beim Aktualisieren von Monat {month}")
                else:
                    # Insert neuen monatlichen Datensatz
                    insert_fields = ['id_empleado', 'anio', 'mes'] + list(filtered_data.keys())
                    insert_values = [employee_id, year, month] + list(filtered_data.values())
                    placeholders = ', '.join(['%s'] * len(insert_fields))
                    query = f"""
                    INSERT INTO t003_ingresos_brutos_mensuales ({', '.join(insert_fields)}) 
                    VALUES ({placeholders})
                    """
                    result = self.execute_update(query, tuple(insert_values))
                    if not result:
                        success_all = False
                        self.logger.error(f"Fehler beim Einfügen von Monat {month}")
            return success_all
        except Exception as e:
            self.logger.error(f"Fehler beim Aktualisieren der Bruttoeinkünfte: {e}")
            return False
    def update_deducciones(self, employee_id: int, year: int, data: Dict[str, Any]) -> bool:
        """Aktualisiert oder fügt Abzüge für einen Mitarbeiter für alle Monate eines Jahres hinzu"""
        try:
            allowed_fields = ['seguro_accidentes', 'adelas', 'sanitas', 'gasolina', 'ret_especie', 'seguro_medico', 'cotizacion_especie']
            # Filtere die Daten auf erlaubte Felder
            filtered_data = {k: v for k, v in data.items() if k in allowed_fields}
            if not filtered_data:
                return False
            # Für jeden Monat (1-12) die monatlichen Datensätze aktualisieren
            success_all = True
            for month in range(1, 13):
                # Prüfen ob monatlicher Datensatz existiert
                check_query = """
                SELECT id_empleado FROM t004_deducciones_mensuales 
                WHERE id_empleado = %s AND anio = %s AND mes = %s
                """
                exists = self.execute_query(check_query, (employee_id, year, month))
                if exists:
                    # Update monatlicher Datensatz
                    update_fields = []
                    update_values = []
                    for field, value in filtered_data.items():
                        update_fields.append(f"{field} = %s")
                        update_values.append(value)
                    query = f"""
                    UPDATE t004_deducciones_mensuales 
                    SET {', '.join(update_fields)} 
                    WHERE id_empleado = %s AND anio = %s AND mes = %s
                    """
                    update_values.extend([employee_id, year, month])
                    result = self.execute_update(query, tuple(update_values))
                    if not result:
                        success_all = False
                        self.logger.error(f"Fehler beim Aktualisieren von Monat {month}")
                else:
                    # Insert neuen monatlichen Datensatz
                    insert_fields = ['id_empleado', 'anio', 'mes'] + list(filtered_data.keys())
                    insert_values = [employee_id, year, month] + list(filtered_data.values())
                    placeholders = ', '.join(['%s'] * len(insert_fields))
                    query = f"""
                    INSERT INTO t004_deducciones_mensuales ({', '.join(insert_fields)}) 
                    VALUES ({placeholders})
                    """
                    result = self.execute_update(query, tuple(insert_values))
                    if not result:
                        success_all = False
                        self.logger.error(f"Fehler beim Einfügen von Monat {month}")
            return success_all
        except Exception as e:
            self.logger.error(f"Fehler beim Aktualisieren der Abzüge: {e}")
            return False
    def get_active_employee_ids(self) -> List[int]:
        try:
            query = "SELECT id_empleado FROM t001_empleados WHERE activo = TRUE ORDER BY id_empleado"
            rows = self.execute_query(query)
            if not rows:
                return []
            return [int(r['id_empleado']) for r in rows if r.get('id_empleado') is not None]
        except Exception as e:
            self.logger.error(f"Fehler beim Laden aktiver Mitarbeiter-IDs: {e}")
            return []
    def get_active_employee_ids_filtered(self, categoria: Optional[str] = None) -> List[int]:
        try:
            category_norm = self._normalize_employee_category(categoria)
            if category_norm is not None and not self._is_valid_employee_category(category_norm):
                return []
            if category_norm is None:
                return self.get_active_employee_ids()
            query = """
            SELECT id_empleado
            FROM t001_empleados
            WHERE activo = TRUE AND categoria = %s
            ORDER BY id_empleado
            """
            rows = self.execute_query(query, (category_norm,))
            if not rows:
                return []
            return [int(r['id_empleado']) for r in rows if r.get('id_empleado') is not None]
        except Exception as e:
            self.logger.error(f"Fehler beim Laden aktiver Mitarbeiter-IDs (gefiltert): {e}")
            return []
    def apply_yearly_ingresos_and_deducciones_to_all_active(
        self,
        year: int,
        ingresos: Optional[Dict[str, Any]] = None,
        deducciones: Optional[Dict[str, Any]] = None,
        categoria: Optional[str] = None,
    ) -> Dict[str, Any]:
        try:
            if ingresos is None and deducciones is None:
                return {
                    "success": False,
                    "message": "Mindestens ingresos oder deducciones muss gesetzt sein",
                    "updated_count": 0,
                    "total_count": 0,
                    "errors": ["Keine Daten übergeben"],
                }
            employee_ids = self.get_active_employee_ids_filtered(categoria=categoria)
            total_count = len(employee_ids)
            if total_count == 0:
                return {
                    "success": True,
                    "message": "Keine aktiven Mitarbeiter gefunden",
                    "updated_count": 0,
                    "total_count": 0,
                    "errors": [],
                }
            updated_count = 0
            errors: List[str] = []
            for employee_id in employee_ids:
                try:
                    ok = True
                    if ingresos is not None:
                        ok = ok and self.update_ingresos(employee_id, year, ingresos)
                    if deducciones is not None:
                        ok = ok and self.update_deducciones(employee_id, year, deducciones)
                    if ok:
                        updated_count += 1
                    else:
                        errors.append(f"Fehler beim Aktualisieren für Mitarbeiter {employee_id}")
                except Exception as e:
                    errors.append(f"Ausnahme bei Mitarbeiter {employee_id}: {str(e)}")
            return {
                "success": updated_count > 0 and len(errors) == 0,
                "message": f"Bulk-Update abgeschlossen: {updated_count} von {total_count} Mitarbeitern aktualisiert",
                "updated_count": updated_count,
                "total_count": total_count,
                "errors": errors,
            }
        except Exception as e:
            self.logger.error(f"Fehler beim Bulk-Update ingresos/deducciones für Jahr {year}: {e}")
            return {
                "success": False,
                "message": "Interner Serverfehler beim Bulk-Update",
                "updated_count": 0,
                "total_count": 0,
                "errors": [str(e)],
            }
    def _update_monthly_from_yearly(self, employee_id: int, year: int, data: Dict[str, Any], table_type: str) -> None:
        """Aktualisiert alle monatlichen Datensätze basierend auf den jährlichen Daten"""
        try:
            if table_type == 'ingresos':
                table_name = 't003_ingresos_brutos_mensuales'
                allowed_fields = [
                     'ticket_restaurant', 'primas', 
                     'dietas_cotizables', 'horas_extras', 'dias_exentos', 
                     'dietas_exentas', 'seguro_pensiones', 'lavado_coche', 'formacion', 'tickets'
                 ]
            elif table_type == 'deducciones':
                table_name = 't004_deducciones_mensuales'
                allowed_fields = ['seguro_accidentes', 'adelas', 'sanitas', 'gasolina', 'ret_especie', 'seguro_medico', 'cotizacion_especie']
            else:
                return
            # Für jeden Monat (1-12) die monatlichen Datensätze aktualisieren
            for month in range(1, 13):
                # Prüfen ob monatlicher Datensatz existiert
                check_query = f"""
                SELECT id_empleado FROM {table_name} 
                WHERE id_empleado = %s AND anio = %s AND mes = %s
                """
                exists = self.execute_query(check_query, (employee_id, year, month))
                if exists:
                    # Update existierenden monatlichen Datensatz
                    update_fields = []
                    update_values = []
                    for field, value in data.items():
                        if field in allowed_fields:
                            update_fields.append(f"{field} = %s")
                            update_values.append(value)
                    if update_fields:
                        query = f"""
                        UPDATE {table_name} 
                        SET {', '.join(update_fields)} 
                        WHERE id_empleado = %s AND anio = %s AND mes = %s
                        """
                        update_values.extend([employee_id, year, month])
                        self.execute_update(query, tuple(update_values))
                else:
                    # Create neuen monatlichen Datensatz
                    insert_fields = ['id_empleado', 'anio', 'mes'] + [field for field in data.keys() if field in allowed_fields]
                    insert_values = [employee_id, year, month] + [data[field] for field in insert_fields[3:]]
                    placeholders = ', '.join(['%s'] * len(insert_fields))
                    query = f"""
                    INSERT INTO {table_name} ({', '.join(insert_fields)}) 
                    VALUES ({placeholders})
                    """
                    self.execute_update(query, tuple(insert_values))
        except Exception as e:
            self.logger.error(f"Fehler bei der Aktualisierung der monatlichen Daten aus jährlichen Daten: {e}")
    def delete_salary(self, employee_id: int, year: int) -> bool:
        """Löscht einen Gehaltsdatensatz für einen Mitarbeiter und Jahr"""
        try:
            query = "DELETE FROM t002_salarios WHERE id_empleado = %s AND anio = %s"
            result = self.execute_update(query, (employee_id, year))
            if result:
                self.logger.info(f"Gehaltsdatensatz für Mitarbeiter {employee_id}, Jahr {year} wurde gelöscht")
            return result
        except Exception as e:  
            self.logger.error(f"Fehler beim Löschen des Gehaltsdatensatzes für Mitarbeiter {employee_id}, Jahr {year}: {e}")
            return False
    def delete_employee(self, employee_id):
        """Löscht einen Mitarbeiter und alle zugehörigen Daten"""
        last_error: Optional[Error] = None
        for attempt in range(2):
            connection = None
            cursor = None
            try:
                connection = self._create_connection()
                cursor = connection.cursor()
                # Zuerst abhängige Daten löschen
                delete_salaries = "DELETE FROM t002_salarios WHERE id_empleado = %s"
                cursor.execute(delete_salaries, (employee_id,))
                delete_ingresos_mensuales = "DELETE FROM t003_ingresos_brutos_mensuales WHERE id_empleado = %s"
                cursor.execute(delete_ingresos_mensuales, (employee_id,))
                delete_deducciones_mensuales = "DELETE FROM t004_deducciones_mensuales WHERE id_empleado = %s"
                cursor.execute(delete_deducciones_mensuales, (employee_id,))
                delete_employee_query = "DELETE FROM t001_empleados WHERE id_empleado = %s"
                cursor.execute(delete_employee_query, (employee_id,))
                connection.commit()
                self.logger.info(f"Mitarbeiter {employee_id} und alle zugehörigen Daten wurden gelöscht")
                return True
            except Error as e:
                last_error = e
                try:
                    if connection is not None:
                        connection.rollback()
                except Exception:
                    pass
                if attempt == 0 and self._is_transient_connection_error(e):
                    self.logger.warning(f"Datenbankverbindung verloren, Retry... ({e})")
                    continue
                self.logger.error(f"Fehler beim Löschen des Mitarbeiters {employee_id}: {e}")
                return False
            except Exception as e:
                try:
                    if connection is not None:
                        connection.rollback()
                except Exception:
                    pass
                self.logger.error(f"Fehler beim Löschen des Mitarbeiters {employee_id}: {e}")
                return False
            finally:
                if cursor is not None:
                    try:
                        cursor.close()
                    except Exception:
                        pass
                if connection is not None:
                    try:
                        connection.close()
                    except Exception:
                        pass
        if last_error is not None:
            self.logger.error(f"Fehler beim Löschen des Mitarbeiters {employee_id}: {last_error}")
        return False
    def hash_password(self, password: str) -> str:
        """Hash a password using SHA-256"""
        return hashlib.sha256(password.encode()).hexdigest()
    def verify_user(self, username: str, password: str) -> Optional[Dict[str, Any]]:
        """Verify user credentials and return user data if valid"""
        try:
            password_hash = self.hash_password(password)
            query = """
            SELECT id_usuario, nombre_usuario, nombre_completo, rol, activo
            FROM t005_usuarios 
            WHERE nombre_usuario = %s AND hash_contraseña = %s AND activo = TRUE
            """
            users = self.execute_query(query, (username, password_hash))
            if users:
                return users[0]
            return None
        except Exception as e:
            self.logger.error(f"Fehler bei der Benutzerüberprüfung: {e}")
            return None
    def create_user(self, username: str, password: str, voller_name: str, rolle: str = 'benutzer') -> bool:
        """Create a new user"""
        try:
            # Check if user already exists
            check_query = "SELECT id_usuario FROM t005_usuarios WHERE nombre_usuario = %s"
            existing = self.execute_query(check_query, (username,))
            if existing:
                self.logger.error(f"Benutzer {username} existiert bereits")
                return False
            # Create new user
            password_hash = self.hash_password(password)
            insert_query = """
            INSERT INTO t005_usuarios (nombre_usuario, hash_contraseña, nombre_completo, rol) 
            VALUES (%s, %s, %s, %s)
            """
            return self.execute_update(insert_query, (username, password_hash, voller_name, rolle))
        except Exception as e:
            self.logger.error(f"Fehler beim Erstellen des Benutzers: {e}")
            return False
    def export_nomina_excel(
        self,
        year: int,
        output_path: str,
        month: int = None,
        extra: bool = False,
    ) -> bool:
        """Exportiert Gehaltsdaten im Excel-Format - nur monatlicher Export wird unterstützt"""
        return DatabaseManagerExportsMixin.export_nomina_excel(self, year, output_path, month, extra=extra)
    def export_asiento_nomina_excel(self, year: int, month: int, output_path: str) -> bool:
        """Exportiert Gehaltsdaten im Asiento Nomina Excel-Format"""
        return DatabaseManagerExportsMixin.export_asiento_nomina_excel(self, year, month, output_path)

    def export_irpf_excel(
        self,
        year: int,
        output_path: str,
        month: int = None,
        extra: bool = False,
    ) -> bool:
        return DatabaseManagerExportsMixin.export_irpf_excel(self, year, month, output_path, extra=extra)
    
    def get_user_email(self, username: str) -> Optional[str]:
        """Holt die Email-Adresse eines Benutzers (angenommen als nombre_usuario)"""
        try:
            query = "SELECT nombre_usuario FROM t005_usuarios WHERE nombre_usuario = %s AND activo = TRUE"
            users = self.execute_query(query, (username,))
            if users:
                # Da nombre_usuario als Email verwendet wird, geben wir diesen zurück
                return users[0]['nombre_usuario']
            return None
        except Exception as e:
            self.logger.error(f"Fehler beim Holen der Benutzer-Email: {e}")
            return None
    
    def create_password_reset_token(self, username: str, email: str, token: str, expires_at: datetime) -> bool:
        """Erstellt einen Passwort-Reset-Token"""
        try:
            # Alte Tokens für diesen Benutzer löschen
            self.delete_password_reset_tokens(username)
            
            query = """
            INSERT INTO t009_password_reset_tokens (nombre_usuario, token, email, expires_at)
            VALUES (%s, %s, %s, %s)
            """
            return self.execute_update(query, (username, token, email, expires_at))
        except Exception as e:
            self.logger.error(f"Fehler beim Erstellen des Passwort-Reset-Tokens: {e}")
            return False
    
    def delete_password_reset_tokens(self, username: str) -> bool:
        """Löscht alle Passwort-Reset-Tokens für einen Benutzer"""
        try:
            query = "DELETE FROM t009_password_reset_tokens WHERE nombre_usuario = %s"
            return self.execute_update(query, (username,))
        except Exception as e:
            self.logger.error(f"Fehler beim Löschen der Passwort-Reset-Tokens: {e}")
            return False
    
    def validate_password_reset_token(self, token: str) -> Optional[Dict[str, Any]]:
        """Validiert einen Passwort-Reset-Token"""
        try:
            query = """
            SELECT nombre_usuario, email, expires_at, used
            FROM t009_password_reset_tokens 
            WHERE token = %s AND used = FALSE
            """
            self.logger.info(f"Validiere Token: {token}")
            tokens = self.execute_query(query, (token,))
            self.logger.info(f"Gefundene Tokens: {tokens}")
            
            if tokens:
                token_data = tokens[0]
                self.logger.info(f"Token gefunden: {token_data}")
                
                # Beide Zeiten auf UTC bringen
                from datetime import datetime, timezone
                now_utc = datetime.now(timezone.utc)
                expires_at = token_data['expires_at']
                
                # Wenn expires_at keine Zeitzone hat, als UTC behandeln
                if expires_at.tzinfo is None:
                    expires_at = expires_at.replace(tzinfo=timezone.utc)
                
                self.logger.info(f"Aktuelle Zeit UTC: {now_utc}")
                self.logger.info(f"Token expires_at: {expires_at}")
                self.logger.info(f"Vergleich: {expires_at} > {now_utc} = {expires_at > now_utc}")
                
                if expires_at > now_utc:
                    self.logger.info(f"Token gültig für: {token_data['nombre_usuario']}")
                    return token_data
                else:
                    self.logger.info(f"Token abgelaufen. Expires: {expires_at}, Now: {now_utc}")
                    return None
            else:
                self.logger.info("Kein gültiger Token gefunden")
                return None
        except Exception as e:
            self.logger.error(f"Fehler bei der Validierung des Passwort-Reset-Tokens: {e}")
            return None
    
    def update_password(self, username: str, new_password: str) -> bool:
        """Aktualisiert das Passwort eines Benutzers"""
        try:
            password_hash = self.hash_password(new_password)
            query = """
            UPDATE t005_usuarios 
            SET hash_contraseña = %s, fecha_modificacion = CURRENT_TIMESTAMP
            WHERE nombre_usuario = %s
            """
            success = self.execute_update(query, (password_hash, username))
            if success:
                # Token als verwendet markieren
                self.mark_token_as_used(username)
            return success
        except Exception as e:
            self.logger.error(f"Fehler beim Aktualisieren des Passworts: {e}")
            return False
    
    def mark_token_as_used(self, username: str) -> bool:
        """Markiert alle Tokens eines Benutzers als verwendet"""
        try:
            query = "UPDATE t009_password_reset_tokens SET used = TRUE WHERE nombre_usuario = %s"
            return self.execute_update(query, (username,))
        except Exception as e:
            self.logger.error(f"Fehler beim Markieren des Tokens als verwendet: {e}")
            return False
    
    def cleanup_expired_tokens(self) -> bool:
        """Löscht abgelaufene Tokens"""
        try:
            query = "DELETE FROM t009_password_reset_tokens WHERE expires_at <= NOW()"
            return self.execute_update(query)
        except Exception as e:
            self.logger.error(f"Fehler beim Aufräumen abgelaufener Tokens: {e}")
            return False
    def update_ingresos_mensuales(self, employee_id: int, year: int, month: int, data: Dict[str, Any]) -> bool:
        """Aktualisiert monatliche Bruttoeinkünfte für einen spezifischen Monat"""
        try:
            allowed_fields = [
                'ticket_restaurant', 'primas', 
                'dietas_cotizables', 'horas_extras', 'dias_exentos', 
                'dietas_exentas', 'seguro_pensiones', 'lavado_coche', 'beca_escolar', 'formacion', 'tickets'
            ]
            # Filtere die Daten auf erlaubte Felder
            filtered_data = {k: v for k, v in data.items() if k in allowed_fields}
            if not filtered_data:
                return False
            # Prüfen ob monatlicher Datensatz existiert
            check_query = """
            SELECT id_empleado FROM t003_ingresos_brutos_mensuales 
            WHERE id_empleado = %s AND anio = %s AND mes = %s
            """
            exists = self.execute_query(check_query, (employee_id, year, month))
            if exists:
                # Update existierenden monatlichen Datensatz
                update_fields = []
                update_values = []
                for field, value in filtered_data.items():
                    update_fields.append(f"{field} = %s")
                    update_values.append(value)
                query = f"""
                UPDATE t003_ingresos_brutos_mensuales 
                SET {', '.join(update_fields)} 
                WHERE id_empleado = %s AND anio = %s AND mes = %s
                """
                update_values.extend([employee_id, year, month])
                return self.execute_update(query, tuple(update_values))
            else:
                # Insert neuen monatlichen Datensatz
                insert_fields = ['id_empleado', 'anio', 'mes'] + list(filtered_data.keys())
                insert_values = [employee_id, year, month] + list(filtered_data.values())
                placeholders = ', '.join(['%s'] * len(insert_fields))
                query = f"""
                INSERT INTO t003_ingresos_brutos_mensuales ({', '.join(insert_fields)}) 
                VALUES ({placeholders})
                """
                return self.execute_update(query, tuple(insert_values))
        except Exception as e:
            self.logger.error(f"Fehler beim Aktualisieren der monatlichen Bruttoeinkünfte: {e}")
            return False
    def update_deducciones_mensuales(self, employee_id: int, year: int, month: int, data: Dict[str, Any]) -> Dict[str, Any]:
        """Aktualisiert monatliche Abzüge für einen spezifischen Monat"""
        result = {
            "success": False,
            "propagation_info": None,
            "error": None
        }
        
        try:
            allowed_fields = ['seguro_accidentes', 'adelas', 'sanitas', 'gasolina', 'ret_especie', 'seguro_medico', 'cotizacion_especie']
            # Filtere die Daten auf erlaubte Felder
            filtered_data = {k: v for k, v in data.items() if k in allowed_fields}
            if not filtered_data:
                result["error"] = "Keine gültigen Felder zum Aktualisieren"
                return result
            
            # Prüfen ob monatlicher Datensatz existiert
            check_query = """
            SELECT id_empleado FROM t004_deducciones_mensuales 
            WHERE id_empleado = %s AND anio = %s AND mes = %s
            """
            exists = self.execute_query(check_query, (employee_id, year, month))
            
            # Alte cotizacion_especie Wert für Vergleich holen
            old_cotizacion_especie = None
            if exists:
                old_query = """
                SELECT cotizacion_especie FROM t004_deducciones_mensuales 
                WHERE id_empleado = %s AND anio = %s AND mes = %s
                """
                old_result = self.execute_query(old_query, (employee_id, year, month))
                if old_result:
                    old_cotizacion_especie = old_result[0].get('cotizacion_especie')
            
            if exists:
                # Update existierenden monatlichen Datensatz
                update_fields = []
                update_values = []
                for field, value in filtered_data.items():
                    update_fields.append(f"{field} = %s")
                    update_values.append(value)
                query = f"""
                UPDATE t004_deducciones_mensuales 
                SET {', '.join(update_fields)} 
                WHERE id_empleado = %s AND anio = %s AND mes = %s
                """
                update_values.extend([employee_id, year, month])
                update_result = self.execute_update(query, tuple(update_values))
            else:
                # Insert neuen monatlichen Datensatz
                insert_fields = ['id_empleado', 'anio', 'mes'] + list(filtered_data.keys())
                insert_values = [employee_id, year, month] + list(filtered_data.values())
                placeholders = ', '.join(['%s'] * len(insert_fields))
                query = f"""
                INSERT INTO t004_deducciones_mensuales ({', '.join(insert_fields)}) 
                VALUES ({placeholders})
                """
                update_result = self.execute_update(query, tuple(insert_values))
            
            result["success"] = update_result
            
            # Automatische Übernahme für cotizacion_especie in folgende Monate
            if 'cotizacion_especie' in filtered_data:
                new_cotizacion_especie = filtered_data['cotizacion_especie']
                
                # Nur ausführen, wenn cotizacion_especie > 0 und sich geändert hat
                if new_cotizacion_especie is not None and new_cotizacion_especie > 0 and old_cotizacion_especie != new_cotizacion_especie:
                    propagation_result = self._propagate_cotizacion_especie(employee_id, year, month, new_cotizacion_especie)
                    result["propagation_info"] = propagation_result
            
            return result
        except Exception as e:
            self.logger.error(f"Fehler beim Aktualisieren der monatlichen Abzüge: {e}")
            result["error"] = str(e)
            return result

    def _propagate_cotizacion_especie(self, employee_id: int, year: int, start_month: int, cotizacion_especie_value: float) -> Dict[str, Any]:
        """Übernimmt cotizacion_especie Wert für alle folgenden Monate bis zum Jahresende"""
        result = {
            "propagated": False,
            "months_updated": [],
            "error": None
        }
        
        try:
            # Alle folgenden Monate bis zum Jahresende aktualisieren
            for month in range(start_month + 1, 13):
                update_query = """
                INSERT INTO t004_deducciones_mensuales (id_empleado, anio, mes, cotizacion_especie)
                VALUES (%s, %s, %s, %s)
                ON DUPLICATE KEY UPDATE 
                    cotizacion_especie = VALUES(cotizacion_especie),
                    fecha_modificacion = CURRENT_TIMESTAMP
                """
                self.execute_update(update_query, (employee_id, year, month, cotizacion_especie_value))
                result["months_updated"].append(month)
            
            if result["months_updated"]:
                result["propagated"] = True
                self.logger.info(f"cotizacion_especie Wert {cotizacion_especie_value} für Mitarbeiter {employee_id} "
                                f"von Monat {start_month} bis Dezember {year} übernommen (Monate: {result['months_updated']})")
            
            return result
        except Exception as e:
            self.logger.error(f"Fehler bei der Übernahme von cotizacion_especie: {e}")
            result["error"] = str(e)
            return result

    def _parse_employee_name_cell(self, raw_value: Any) -> Optional[Dict[str, str]]:
        try:
            if raw_value is None:
                return None
            text = str(raw_value).strip()
            if not text:
                return None
            # Expected: "APELLIDO, NOMBRE" (user mentioned NACHNAME, VORNAME)
            parts = [p.strip() for p in text.split(',')]
            if len(parts) < 2:
                return None
            apellido = parts[0]
            nombre = ','.join(parts[1:]).strip()
            if not apellido or not nombre:
                return None
            return {"apellido": apellido, "nombre": nombre}
        except Exception:
            return None
    def _to_decimal_number(self, raw_value: Any) -> float:
        if raw_value is None:
            return 0.0
        if isinstance(raw_value, (int, float)):
            try:
                return float(raw_value)
            except Exception:
                return 0.0
        try:
            s = str(raw_value).strip()
            if not s:
                return 0.0
            s = s.replace('.', '').replace(',', '.') if s.count(',') == 1 and s.count('.') >= 1 else s
            s = s.replace(',', '.')
            return float(s)
        except Exception:
            return 0.0
    def _find_employee_id_by_name(self, apellido: str, nombre: str) -> Optional[int]:
        try:
            query = """
            SELECT id_empleado
            FROM t001_empleados
            WHERE LOWER(TRIM(apellido)) = LOWER(TRIM(%s))
              AND LOWER(TRIM(nombre)) = LOWER(TRIM(%s))
            LIMIT 1
            """
            rows = self.execute_query(query, (apellido, nombre))
            if rows:
                return int(rows[0]["id_empleado"])
            return None
        except Exception as e:
            self.logger.error(f"Fehler beim Suchen empleado_id für {apellido}, {nombre}: {e}")
            return None
    def import_horas_dietas_worksheet(
        self,
        worksheet: Worksheet,
        year: int,
        month: int,
        usuario_login: str,
        source_filename: str,
    ) -> Dict[str, Any]:
        """Importiert Horas + Dietas Daten in t003_ingresos_brutos_mensuales (monatlich)."""
        errors: List[str] = []
        row_results: List[Dict[str, Any]] = []
        processed_count = 0
        updated_count = 0
        inserted_count = 0
        skipped_count = 0
        if worksheet is None:
            return {"success": False, "message": "worksheet fehlt"}
        try:
            # Preload employee lookup map to avoid many DB queries
            employees = self.execute_query("SELECT id_empleado, nombre, apellido FROM t001_empleados") or []
            employee_map: Dict[str, int] = {}
            for e in employees:
                key = f"{str(e.get('apellido','')).strip().lower()}|{str(e.get('nombre','')).strip().lower()}"
                if key and e.get('id_empleado') is not None:
                    employee_map[key] = int(e['id_empleado'])
            def get_employee_id(apellido: str, nombre: str) -> Optional[int]:
                key = f"{apellido.strip().lower()}|{nombre.strip().lower()}"
                if key in employee_map:
                    return employee_map[key]
                # fallback DB query (in case of late insert)
                emp_id = self._find_employee_id_by_name(apellido, nombre)
                if emp_id is not None:
                    employee_map[key] = emp_id
                return emp_id
            # Iterate rows, data starts usually at row 2
            max_row = worksheet.max_row or 0
            for row_idx in range(2, max_row + 1):
                try:
                    name_cell = worksheet[f"A{row_idx}"].value
                    parsed = self._parse_employee_name_cell(name_cell)
                    if not parsed:
                        # empty row - skip silently
                        skipped_count += 1
                        continue
                    apellido = parsed['apellido']
                    nombre = parsed['nombre']
                    employee_id = get_employee_id(apellido, nombre)
                    if employee_id is None:
                        msg = f"Zeile {row_idx}: Mitarbeiter nicht gefunden: {apellido}, {nombre}"
                        errors.append(msg)
                        row_results.append({
                            "row": row_idx,
                            "employee": f"{apellido}, {nombre}",
                            "success": False,
                            "error": "employee_not_found",
                        })
                        continue
                    horas_extras = self._to_decimal_number(worksheet[f"F{row_idx}"].value)
                    primas_g = self._to_decimal_number(worksheet[f"G{row_idx}"].value)
                    primas_h = self._to_decimal_number(worksheet[f"H{row_idx}"].value)
                    primas = primas_g + primas_h
                    dias_exentos = self._to_decimal_number(worksheet[f"I{row_idx}"].value)
                    dietas_cotizables = self._to_decimal_number(worksheet[f"J{row_idx}"].value)
                    dietas_exentas = self._to_decimal_number(worksheet[f"K{row_idx}"].value)
                    payload = {
                        "horas_extras": horas_extras,
                        "primas": primas,
                        "dias_exentos": dias_exentos,
                        "dietas_cotizables": dietas_cotizables,
                        "dietas_exentas": dietas_exentas,
                    }
                    # Determine insert vs update
                    check_query = """
                    SELECT id_empleado
                    FROM t003_ingresos_brutos_mensuales
                    WHERE id_empleado = %s AND anio = %s AND mes = %s
                    """
                    exists = self.execute_query(check_query, (employee_id, year, month))
                    old_data = None
                    if exists:
                        old_rows = self.execute_query(
                            """
                            SELECT ticket_restaurant, primas, dietas_cotizables, horas_extras, dias_exentos, dietas_exentas,
                                   seguro_pensiones, lavado_coche, formacion, tickets
                            FROM t003_ingresos_brutos_mensuales
                            WHERE id_empleado = %s AND anio = %s AND mes = %s
                            """,
                            (employee_id, year, month),
                        )
                        old_data = old_rows[0] if old_rows else None
                    success = self.update_ingresos_mensuales(employee_id, year, month, payload)
                    if not success:
                        msg = f"Zeile {row_idx}: Fehler beim Speichern für empleado_id={employee_id}"
                        errors.append(msg)
                        row_results.append({
                            "row": row_idx,
                            "employee": f"{apellido}, {nombre}",
                            "employee_id": employee_id,
                            "success": False,
                            "error": "db_write_failed",
                        })
                        continue
                    processed_count += 1
                    if exists:
                        updated_count += 1
                    else:
                        inserted_count += 1
                    try:
                        # Bearbeitungslog: use diff when possible
                        new_data = payload
                        details = None
                        try:
                            details = self.create_change_details(old_data=old_data, new_data=new_data)
                        except Exception:
                            details = {
                                "source": "import_horas_dietas",
                                "filename": source_filename,
                                "row": row_idx,
                                "data": new_data,
                            }
                        self.insert_registro_procesamiento(
                            usuario_login=usuario_login,
                            accion="import",
                            objeto="ingresos_mensuales",
                            id_empleado=employee_id,
                            anio=year,
                            mes=month,
                            details=details,
                        )
                    except Exception:
                        pass
                    row_results.append({
                        "row": row_idx,
                        "employee": f"{apellido}, {nombre}",
                        "employee_id": employee_id,
                        "success": True,
                        "data": payload,
                    })
                except Exception as e:
                    msg = f"Zeile {row_idx}: Unerwarteter Fehler: {e}"
                    errors.append(msg)
                    row_results.append({"row": row_idx, "success": False, "error": "unexpected"})
            success_overall = processed_count > 0 and len(errors) == 0
            return {
                "success": success_overall,
                "year": year,
                "month": month,
                "processed_count": processed_count,
                "inserted_count": inserted_count,
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "error_count": len(errors),
                "errors": errors,
                "rows": row_results,
                "message": f"Import abgeschlossen: {processed_count} verarbeitet ({inserted_count} neu, {updated_count} aktualisiert), {len(errors)} Fehler.",
            }
        except Exception as e:
            self.logger.error(f"Fehler beim Import Horas+Dietas Worksheet: {e}")
            return {"success": False, "message": f"Import fehlgeschlagen: {str(e)}"}

    def import_gasolina_worksheet(
        self,
        worksheet: Worksheet,
        year: int,
        month: int,
        usuario_login: str,
        source_filename: str,
    ) -> Dict[str, Any]:
        """Importiert Gasolina (monatlich) in t004_deducciones_mensuales.gasolina.

        Expected format:
        - Column A: employee name in format "APELLIDO, NOMBRE"
        - Column B: amount
        """
        errors: List[str] = []
        row_results: List[Dict[str, Any]] = []
        processed_count = 0
        updated_count = 0
        inserted_count = 0
        skipped_count = 0
        if worksheet is None:
            return {"success": False, "message": "worksheet fehlt"}
        try:
            employees = self.execute_query("SELECT id_empleado, nombre, apellido FROM t001_empleados") or []
            employee_map: Dict[str, int] = {}
            for e in employees:
                key = f"{str(e.get('apellido','')).strip().lower()}|{str(e.get('nombre','')).strip().lower()}"
                if key and e.get('id_empleado') is not None:
                    employee_map[key] = int(e['id_empleado'])

            def get_employee_id(apellido: str, nombre: str) -> Optional[int]:
                key = f"{apellido.strip().lower()}|{nombre.strip().lower()}"
                if key in employee_map:
                    return employee_map[key]
                emp_id = self._find_employee_id_by_name(apellido, nombre)
                if emp_id is not None:
                    employee_map[key] = emp_id
                return emp_id

            max_row = worksheet.max_row or 0
            for row_idx in range(2, max_row + 1):
                try:
                    name_cell = worksheet[f"A{row_idx}"].value
                    parsed = self._parse_employee_name_cell(name_cell)
                    if not parsed:
                        skipped_count += 1
                        continue

                    apellido = parsed['apellido']
                    nombre = parsed['nombre']
                    employee_id = get_employee_id(apellido, nombre)
                    if employee_id is None:
                        msg = f"Zeile {row_idx}: Mitarbeiter nicht gefunden: {apellido}, {nombre}"
                        errors.append(msg)
                        row_results.append({
                            "row": row_idx,
                            "employee": f"{apellido}, {nombre}",
                            "success": False,
                            "error": "employee_not_found",
                        })
                        continue

                    gasolina = self._to_decimal_number(worksheet[f"B{row_idx}"].value)
                    
                    # Überspringen, wenn kein gültiger Wert (0.0 oder None)
                    if gasolina == 0.0:
                        skipped_count += 1
                        continue
                        
                    payload = {"gasolina": gasolina}

                    check_query = """
                    SELECT id_empleado
                    FROM t004_deducciones_mensuales
                    WHERE id_empleado = %s AND anio = %s AND mes = %s
                    """
                    exists = self.execute_query(check_query, (employee_id, year, month))
                    old_data = None
                    if exists:
                        old_rows = self.execute_query(
                            """
                            SELECT seguro_accidentes, adelas, sanitas, gasolina, ret_especie, seguro_medico, cotizacion_especie
                            FROM t004_deducciones_mensuales
                            WHERE id_empleado = %s AND anio = %s AND mes = %s
                            """,
                            (employee_id, year, month),
                        )
                        old_data = old_rows[0] if old_rows else None

                    result = self.update_deducciones_mensuales(employee_id, year, month, payload)
                    if not result.get("success", False):
                        msg = f"Zeile {row_idx}: Fehler beim Speichern für empleado_id={employee_id}"
                        if result.get("error"):
                            msg += f": {result['error']}"
                        errors.append(msg)
                        row_results.append({
                            "row": row_idx,
                            "employee": f"{apellido}, {nombre}",
                            "employee_id": employee_id,
                            "success": False,
                            "error": "db_write_failed",
                            "details": result.get("error"),
                        })
                        continue

                    processed_count += 1
                    if exists:
                        updated_count += 1
                    else:
                        inserted_count += 1

                    try:
                        new_data = payload
                        details = None
                        try:
                            details = self.create_change_details(old_data=old_data, new_data=new_data)
                        except Exception:
                            details = {
                                "source": "import_gasolina",
                                "filename": source_filename,
                                "row": row_idx,
                                "data": new_data,
                            }
                        self.insert_registro_procesamiento(
                            usuario_login=usuario_login,
                            accion="import",
                            objeto="deducciones_mensuales",
                            id_empleado=employee_id,
                            anio=year,
                            mes=month,
                            details=details,
                        )
                    except Exception:
                        pass

                    row_results.append({
                        "row": row_idx,
                        "employee": f"{apellido}, {nombre}",
                        "employee_id": employee_id,
                        "success": True,
                        "data": payload,
                    })
                except Exception as e:
                    msg = f"Zeile {row_idx}: Unerwarteter Fehler: {e}"
                    errors.append(msg)
                    row_results.append({"row": row_idx, "success": False, "error": "unexpected"})

            # Erfolgsmeldung anpassen - Import ist erfolgreich wenn mindestens ein Datensatz verarbeitet wurde,
            # auch wenn einige Zeilen übersprungen wurden (z.B. Mitarbeiter nicht gefunden)
            success_overall = processed_count > 0
            
            # Bessere Erfolgsmeldung basierend auf dem Ergebnis
            if processed_count > 0:
                if len(errors) == 0:
                    message = f"Import erfolgreich abgeschlossen: {processed_count} Datensätze verarbeitet ({inserted_count} neu, {updated_count} aktualisiert)"
                else:
                    message = f"Import teilweise erfolgreich: {processed_count} Datensätze verarbeitet ({inserted_count} neu, {updated_count} aktualisiert), {len(errors)} Zeilen mit Fehlern übersprungen"
            else:
                message = f"Import abgeschlossen: Keine gültigen Datensätze zum Verarbeiten gefunden. {len(errors)} Fehler aufgetreten."
            
            return {
                "success": success_overall,
                "year": year,
                "month": month,
                "processed_count": processed_count,
                "inserted_count": inserted_count,
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "error_count": len(errors),
                "errors": errors,
                "rows": row_results,
                "message": message,
            }
        except Exception as e:
            self.logger.error(f"Fehler beim Import Gasolina Worksheet: {e}")
            return {"success": False, "message": f"Import fehlgeschlagen: {str(e)}"}


    def import_cotizacion_especie_worksheet(
        self,
        worksheet: Worksheet,
        year: int,
        month: int,
        usuario_login: str,
        source_filename: str,
    ) -> Dict[str, Any]:
        """Importiert Cotizacion Especie (monatlich) in t004_deducciones_mensuales.cotizacion_especie.

        Expected format:
        - Column A: employee name in format "APELLIDO, NOMBRE"
        - Column F: amount

        Rows without a matching employee are skipped silently.
        """
        errors: List[str] = []
        row_results: List[Dict[str, Any]] = []
        processed_count = 0
        updated_count = 0
        inserted_count = 0
        skipped_count = 0

        if worksheet is None:
            return {"success": False, "message": "worksheet fehlt"}

        try:
            employees = self.execute_query("SELECT id_empleado, nombre, apellido FROM t001_empleados") or []
            employee_map: Dict[str, int] = {}
            for e in employees:
                key = f"{str(e.get('apellido','')).strip().lower()}|{str(e.get('nombre','')).strip().lower()}"
                if key and e.get('id_empleado') is not None:
                    employee_map[key] = int(e['id_empleado'])

            def get_employee_id(apellido: str, nombre: str) -> Optional[int]:
                key = f"{apellido.strip().lower()}|{nombre.strip().lower()}"
                if key in employee_map:
                    return employee_map[key]
                emp_id = self._find_employee_id_by_name(apellido, nombre)
                if emp_id is not None:
                    employee_map[key] = emp_id
                return emp_id

            max_row = worksheet.max_row or 0
            for row_idx in range(2, max_row + 1):
                try:
                    name_cell = worksheet[f"A{row_idx}"].value
                    parsed = self._parse_employee_name_cell(name_cell)
                    if not parsed:
                        skipped_count += 1
                        continue

                    apellido = parsed['apellido']
                    nombre = parsed['nombre']
                    employee_id = get_employee_id(apellido, nombre)
                    if employee_id is None:
                        skipped_count += 1
                        continue

                    cotizacion_especie = self._to_decimal_number(worksheet[f"F{row_idx}"].value)
                    payload = {"cotizacion_especie": cotizacion_especie}

                    check_query = """
                    SELECT id_empleado
                    FROM t004_deducciones_mensuales
                    WHERE id_empleado = %s AND anio = %s AND mes = %s
                    """
                    exists = self.execute_query(check_query, (employee_id, year, month))
                    old_data = None
                    if exists:
                        old_rows = self.execute_query(
                            """
                            SELECT seguro_accidentes, adelas, sanitas, gasolina, ret_especie, seguro_medico, cotizacion_especie
                            FROM t004_deducciones_mensuales
                            WHERE id_empleado = %s AND anio = %s AND mes = %s
                            """,
                            (employee_id, year, month),
                        )
                        old_data = old_rows[0] if old_rows else None

                    success = self.update_deducciones_mensuales(employee_id, year, month, payload)
                    if not success:
                        msg = f"Zeile {row_idx}: Fehler beim Speichern für empleado_id={employee_id}"
                        errors.append(msg)
                        row_results.append({
                            "row": row_idx,
                            "employee": f"{apellido}, {nombre}",
                            "employee_id": employee_id,
                            "success": False,
                            "error": "db_write_failed",
                        })
                        continue

                    processed_count += 1
                    if exists:
                        updated_count += 1
                    else:
                        inserted_count += 1

                    try:
                        new_data = payload
                        details = None
                        try:
                            details = self.create_change_details(old_data=old_data, new_data=new_data)
                        except Exception:
                            details = {
                                "source": "import_cotizacion_especie",
                                "filename": source_filename,
                                "row": row_idx,
                                "data": new_data,
                            }
                        self.insert_registro_procesamiento(
                            usuario_login=usuario_login,
                            accion="import",
                            objeto="deducciones_mensuales",
                            id_empleado=employee_id,
                            anio=year,
                            mes=month,
                            details=details,
                        )
                    except Exception:
                        pass

                    row_results.append({
                        "row": row_idx,
                        "employee": f"{apellido}, {nombre}",
                        "employee_id": employee_id,
                        "success": True,
                        "data": payload,
                    })

                except Exception as e:
                    msg = f"Zeile {row_idx}: Unerwarteter Fehler: {e}"
                    errors.append(msg)
                    row_results.append({"row": row_idx, "success": False, "error": "unexpected"})

            success_overall = len(errors) == 0
            return {
                "success": success_overall,
                "year": year,
                "month": month,
                "processed_count": processed_count,
                "inserted_count": inserted_count,
                "updated_count": updated_count,
                "skipped_count": skipped_count,
                "error_count": len(errors),
                "errors": errors,
                "rows": row_results,
                "message": f"Import abgeschlossen: {processed_count} verarbeitet ({inserted_count} neu, {updated_count} aktualisiert), {len(errors)} Fehler.",
            }

        except Exception as e:
            self.logger.error(f"Fehler beim Import Cotizacion Especie Worksheet: {e}")
            return {"success": False, "message": f"Import fehlgeschlagen: {str(e)}"}
    def copy_salaries_to_new_year(self, target_year: int) -> Dict[str, Any]:
        """Kopiert Gehälter aller aktiven Mitarbeiter vom Vorjahr ins Zieljahr"""
        try:
            source_year = target_year - 1
            # Prüfen ob Zieljahr in der Zukunft liegt (nicht mehr als 20 Jahre voraus)
            current_year = datetime.now().year
            if target_year > current_year + 20:
                return {
                    "success": False,
                    "message": f"Zieljahr {target_year} ist zu weit in der Zukunft (max. {current_year + 20})",
                    "copied_count": 0,
                    "skipped_count": 0,
                    "errors": []
                }
            # Finde alle aktiven Mitarbeiter mit Gehalt im Vorjahr
            # Für zukünftige Jahre kopiere alle aktiven Mitarbeiter, die Gehalt im Vorjahr haben
            if target_year > datetime.now().year:
                # Zukünftiges Jahr: kopiere alle aktiven Mitarbeiter mit Vorjahresgehalt
                employees_with_salaries_query = """
                SELECT DISTINCT e.id_empleado, e.nombre, e.apellido,
                       s.modalidad, s.antiguedad, s.salario_anual_bruto
                FROM t001_empleados e
                INNER JOIN t002_salarios s ON e.id_empleado = s.id_empleado
                WHERE e.activo = TRUE 
                  AND s.anio = %s
                  AND e.id_empleado NOT IN (
                      SELECT id_empleado FROM t002_salarios WHERE anio = %s
                  )
                ORDER BY e.id_empleado
                """
            else:
                # Vergangenheit/Gegenwart: nur fehlende Mitarbeiter kopieren
                employees_with_salaries_query = """
                SELECT DISTINCT e.id_empleado, e.nombre, e.apellido,
                       s.modalidad, s.antiguedad, s.salario_anual_bruto
                FROM t001_empleados e
                INNER JOIN t002_salarios s ON e.id_empleado = s.id_empleado
                WHERE e.activo = TRUE 
                  AND s.anio = %s
                  AND e.id_empleado NOT IN (
                      SELECT id_empleado FROM t002_salarios WHERE anio = %s
                  )
                ORDER BY e.id_empleado
                """
            employees_to_copy = self.execute_query(employees_with_salaries_query, (source_year, target_year))
            if not employees_to_copy:
                return {
                    "success": True,
                    "message": f"Keine Mitarbeiter gefunden, deren Gehalt für {target_year} kopiert werden muss",
                    "copied_count": 0,
                    "skipped_count": 0,
                    "errors": []
                }
            copied_count = 0
            skipped_count = 0
            errors = []
            for employee in employees_to_copy:
                try:
                    # Gehalt für neues Jahr einfügen
                    salary_data = {
                        'anio': target_year,
                        'modalidad': employee['modalidad'],
                        'antiguedad': employee['antiguedad'],
                        'salario_anual_bruto': employee['salario_anual_bruto']
                    }
                    success = self.add_salary(employee['id_empleado'], salary_data)
                    if success:
                        copied_count += 1
                        self.logger.info(f"Gehalt für Mitarbeiter {employee['id_empleado']} ({employee['nombre']} {employee['apellido']}) nach {target_year} kopiert")
                    else:
                        skipped_count += 1
                        errors.append(f"Fehler beim Kopieren für Mitarbeiter {employee['id_empleado']} ({employee['nombre']} {employee['apellido']})")
                except Exception as e:
                    skipped_count += 1
                    error_msg = f"Ausnahme beim Kopieren für Mitarbeiter {employee['id_empleado']}: {str(e)}"
                    errors.append(error_msg)
                    self.logger.error(error_msg)
            result_message = f"Gehaltskopierung für {target_year} abgeschlossen: {copied_count} kopiert, {skipped_count} übersprungen"
            return {
                "success": copied_count > 0,
                "message": result_message,
                "copied_count": copied_count,
                "skipped_count": skipped_count,
                "errors": errors
            }
        except Exception as e:
            self.logger.error(f"Fehler bei der Gehaltskopierung für Jahr {target_year}: {e}")
            return {
                "success": False,
                "message": f"Allgemeiner Fehler bei der Gehaltskopierung: {str(e)}",
                "copied_count": 0,
                "skipped_count": 0,
                "errors": [str(e)]
            }
    def get_missing_salary_years(self) -> List[Dict[str, Any]]:
        """Gibt eine Liste der Jahre zurück, für die aktive Mitarbeiter keine Gehälter haben, aber nur wenn Vorjahresdaten existieren"""
        try:
            current_year = datetime.now().year
            missing_years = []
            # Prüfe die letzten 2 Jahre und die nächsten 20 Jahre
            for year in range(current_year - 1, current_year + 21):
                # Prüfe zuerst ob es überhaupt Gehaltsdaten im Vorjahr gibt
                prev_year_exists_query = """
                SELECT COUNT(*) as count FROM t002_salarios 
                WHERE anio = %s
                """
                prev_year_result = self.execute_query(prev_year_exists_query, (year - 1,))
                # Wenn keine Vorjahresdaten existieren oder DB-Verbindung fehlt, überspringe dieses Jahr
                if not prev_year_result or prev_year_result[0]['count'] == 0:
                    continue
                # Finde aktive Mitarbeiter ohne Gehalt für dieses Jahr
                missing_query = """
                SELECT e.id_empleado, e.nombre, e.apellido
                FROM t001_empleados e
                WHERE e.activo = TRUE
                  AND e.id_empleado NOT IN (
                      SELECT id_empleado FROM t002_salarios WHERE anio = %s
                  )
                ORDER BY e.id_empleado
                """
                missing_employees = self.execute_query(missing_query, (year,))
                # Zeige Jahre an, wenn:
                # 1. Mitarbeiter fehlen (für Vergangenheit/Gegenwart)
                # 2. Für zukünftige Jahre immer anzeigen (wenn Vorjahresdaten existieren)
                if missing_employees or year > current_year:
                    missing_years.append({
                        "year": year,
                        "missing_count": len(missing_employees),
                        "employees": missing_employees,
                        "is_future_year": year > current_year,
                        "has_previous_year_data": True
                    })
            return missing_years
        except Exception as e:
            self.logger.error(f"Fehler bei der Ermittlung fehlender Gehaltsjahre: {e}")
            return []
    def recalculate_all_atrasos_for_year(self, year: int) -> Dict[str, Any]:
        """
        Berechnet alle Atrasos für alle Mitarbeiter im angegebenen Jahr neu.
        Args:
            year: Jahr für das die Atrasos neu berechnet werden sollen
        Returns:
            Dict mit Erfolgsmeldung und Statistik
        """
        try:
            # Hole alle Gehälter für das angegebene Jahr
            salaries_query = """
            SELECT id_empleado, anio, modalidad, salario_anual_bruto
            FROM t002_salarios
            WHERE anio = %s
            ORDER BY id_empleado
            """
            salaries = self.execute_query(salaries_query, (year,))
            if not salaries:
                return {
                    "success": False,
                    "message": f"Keine Gehälter für Jahr {year} gefunden",
                    "updated_count": 0,
                    "errors": []
                }
            updated_count = 0
            errors = []
            for salary in salaries:
                try:
                    employee_id = salary['id_empleado']
                    modalidad = salary['modalidad']
                    current_salary = float(salary['salario_anual_bruto'])
                    # Berechne neue Atrasos
                    new_atrasos = self.calculate_atrasos(employee_id, year, modalidad, current_salary)
                    new_salary_con_atrasos = current_salary / (modalidad if modalidad in [12, 14] else 12) + new_atrasos
                    # Update der Datenbank
                    update_query = """
                    UPDATE t002_salarios 
                    SET atrasos = %s, 
                        salario_mensual_con_atrasos = %s,
                        fecha_modificacion = CURRENT_TIMESTAMP
                    WHERE id_empleado = %s AND anio = %s
                    """
                    success = self.execute_update(update_query, (
                        new_atrasos, 
                        new_salary_con_atrasos, 
                        employee_id, 
                        year
                    ))
                    if success:
                        updated_count += 1
                        self.logger.info(f"Atrasos für Mitarbeiter {employee_id} in Jahr {year} neu berechnet: {new_atrasos}")
                    else:
                        errors.append(f"Fehler beim Update für Mitarbeiter {employee_id}")
                except Exception as e:
                    error_msg = f"Fehler bei der Neuberechnung für Mitarbeiter {salary['id_empleado']}: {str(e)}"
                    errors.append(error_msg)
                    self.logger.error(error_msg)
            result_message = f"Neuberechnung der Atrasos für {year} abgeschlossen: {updated_count} Mitarbeiter aktualisiert"
            return {
                "success": updated_count > 0,
                "message": result_message,
                "updated_count": updated_count,
                "total_count": len(salaries),
                "errors": errors
            }
        except Exception as e:
            self.logger.error(f"Fehler bei der Neuberechnung der Atrasos für Jahr {year}: {e}")
            return {
                "success": False,
                "message": f"Allgemeiner Fehler bei der Neuberechnung: {str(e)}",
                "updated_count": 0,
                "total_count": 0,
                "errors": [str(e)]
            }
