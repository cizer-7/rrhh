"""
Flask API Server für Mitarbeiter Gehaltsabrechnung
Einfache Alternative zu FastAPI ohne Pydantic-Probleme
"""

from flask import Flask, request, jsonify, send_file
from flask_cors import CORS
import logging
from datetime import datetime, timedelta, timezone
import jwt
import json
import os
import io
from openpyxl import load_workbook
from database_manager import DatabaseManager
from config.email_settings import email_service

# Logging konfigurieren
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

# Flask App
app = Flask(__name__)
CORS(app, origins=[
    "http://localhost:3000", 
    "http://127.0.0.1:3000",
    "https://digitalisierung-frontend.azurewebsites.net"
])

# Security
SECRET_KEY = os.getenv("SECRET_KEY", "dein-geheimer-schlüssel-hier-ändern")
ALGORITHM = "HS256"

def _env_bool(name: str, default: bool = False) -> bool:
    raw = os.getenv(name)
    if raw is None:
        return default
    v = str(raw).strip().lower()
    if v in {"1", "true", "yes", "y", "on"}:
        return True
    if v in {"0", "false", "no", "n", "off"}:
        return False
    return default

# Database Manager Initialisierung
db_manager = DatabaseManager(
    host=os.getenv("DB_HOST", "localhost"),
    database=os.getenv("DB_NAME", "nomina"),
    user=os.getenv("DB_USER", "root"),
    password=os.getenv("DB_PASSWORD", ""),
    port=int(os.getenv("DB_PORT", "3307")),
    ssl_disabled=_env_bool("DB_SSL_DISABLED", default=False),
)

# JWT Token Funktionen
def create_access_token(data: dict, expires_delta: timedelta = None):
    to_encode = data.copy()
    if expires_delta:
        expire = datetime.now(timezone.utc) + expires_delta
    else:
        expire = datetime.now(timezone.utc) + timedelta(hours=24)
    to_encode.update({"exp": expire})
    encoded_jwt = jwt.encode(to_encode, SECRET_KEY, algorithm=ALGORITHM)
    return encoded_jwt

def verify_token(token: str):
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username: str = payload.get("sub")
        if username is None:
            return None
        return username
    except jwt.PyJWTError:
        return None

def token_required(f):
    from functools import wraps
    @wraps(f)
    def decorated(*args, **kwargs):
        token = None
        if 'Authorization' in request.headers:
            auth_header = request.headers['Authorization']
            try:
                token = auth_header.split(" ")[1]  # Bearer <token>
            except IndexError:
                return jsonify({"message": "Token format invalid"}), 401
        
        if not token:
            return jsonify({"message": "Token is missing"}), 401
        
        current_user = verify_token(token)
        if not current_user:
            return jsonify({"message": "Token is invalid"}), 401
        
        return f(current_user, *args, **kwargs)
    return decorated

# Datenbankverbindung herstellen
def setup_database():
    if not db_manager.connect():
        logger.error("Konnte nicht zur Datenbank verbinden!")
        raise Exception("Datenbankverbindung fehlgeschlagen")
    logger.info("API Server gestartet und Datenbank verbunden")

# Initialisierung
setup_database()

@app.before_request
def before_request():
    """Stelle sicher, dass die Datenbankverbindung aktiv ist"""
    try:
        if not db_manager.connection:
            db_manager.connect()
    except Exception as e:
        print(f"Datenbankverbindungsfehler: {e}")
        db_manager.connect()

@app.teardown_appcontext
def shutdown_database(exception=None):
    pass  # Nicht bei jedem Request disconnecten

# Authentifizierung Endpunkte
@app.route('/auth/login', methods=['POST'])
def login():
    """Benutzeranmeldung"""
    try:
        data = request.get_json()
        username = data.get("username")
        password = data.get("password")
        
        logger.info(f"Login-Versuch für Benutzer: {username}")
        
        if not username or not password:
            return jsonify({"error": "Username und Password erforderlich"}), 400
        
        user_data = db_manager.verify_user(username, password)
        logger.info(f"verify_user Ergebnis: {user_data}")
        
        if not user_data:
            return jsonify({"error": "Falscher Benutzername oder Passwort"}), 401
        
        access_token = create_access_token(data={"sub": user_data["nombre_usuario"]})
        return jsonify({
            "access_token": access_token,
            "token_type": "bearer",
            "user": user_data
        })
    except Exception as e:
        logger.error(f"Login Fehler: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({"error": "Interner Serverfehler"}), 500

# Passwort-Reset Endpunkte
@app.route('/auth/forgot-password', methods=['POST'])
def forgot_password():
    """Passwort-Reset-Anfrage"""
    try:
        data = request.get_json()
        username = data.get("username")
        
        if not username:
            return jsonify({"error": "Benutzername erforderlich"}), 400
        
        # Prüfen ob Benutzer existiert
        email = db_manager.get_user_email(username)
        if not email:
            # Sicherheit: Nicht verraten, ob Benutzer existiert
            return jsonify({"message": "Wenn der Benutzer existiert, wurde eine Reset-Email gesendet"}), 200
        
        # Token generieren
        token = email_service.generate_reset_token()
        expires_at = datetime.now(timezone.utc) + timedelta(hours=24)  # 24 Stunden statt 1 Stunde
        
        logger.info(f"Token generiert für {username}: {token}")
        logger.info(f"Token expires at: {expires_at}")
        
        # Token in Datenbank speichern
        success = db_manager.create_password_reset_token(username, email, token, expires_at)
        logger.info(f"Token in DB gespeichert: {success}")
        
        if not success:
            return jsonify({"error": "Fehler beim Erstellen des Reset-Tokens"}), 500
        
        # Email senden
        email_sent = email_service.send_password_reset_email(email, username, token, os.getenv('FRONTEND_URL', 'http://localhost:3000'))
        logger.info(f"Email gesendet: {email_sent}")
        
        if not email_sent:
            # Token trotzdem speichern, aber Fehler melden
            logger.error("Email-Versand fehlgeschlagen, aber Token wurde gespeichert")
            # return jsonify({"error": "Fehler beim Senden der Reset-Email"}), 500
        
        logger.info(f"Passwort-Reset-Token erstellt für Benutzer: {username}")
        return jsonify({"message": "Wenn der Benutzer existiert, wurde eine Reset-Email gesendet"}), 200
        
    except Exception as e:
        logger.error(f"Fehler bei der Passwort-Reset-Anfrage: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/auth/reset-password', methods=['POST'])
def reset_password():
    """Passwort-Reset-Bestätigung"""
    try:
        data = request.get_json()
        token = data.get("token")
        new_password = data.get("new_password")
        
        if not token or not new_password:
            return jsonify({"error": "Token und neues Passwort erforderlich"}), 400
        
        if len(new_password) < 6:
            return jsonify({"error": "Passwort muss mindestens 6 Zeichen lang sein"}), 400
        
        # Token validieren
        token_data = db_manager.validate_password_reset_token(token)
        if not token_data:
            return jsonify({"error": "Ungültiger oder abgelaufener Token"}), 400
        
        username = token_data['nombre_usuario']
        
        # Passwort aktualisieren
        success = db_manager.update_password(username, new_password)
        if not success:
            return jsonify({"error": "Fehler beim Aktualisieren des Passworts"}), 500
        
        logger.info(f"Passwort erfolgreich zurückgesetzt für Benutzer: {username}")
        return jsonify({"message": "Passwort erfolgreich aktualisiert"}), 200
        
    except Exception as e:
        logger.error(f"Fehler beim Passwort-Reset: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/auth/validate-reset-token', methods=['POST'])
def validate_reset_token():
    """Validiert einen Passwort-Reset-Token"""
    try:
        data = request.get_json()
        token = data.get("token")
        
        logger.info(f"Token-Validierung angefordert: {token}")
        
        if not token:
            return jsonify({"error": "Token erforderlich"}), 400
        
        token_data = db_manager.validate_password_reset_token(token)
        logger.info(f"Token-Validierung Ergebnis: {token_data}")
        
        if not token_data:
            return jsonify({"error": "Ungültiger oder abgelaufener Token"}), 400
        
        return jsonify({
            "valid": True,
            "username": token_data['nombre_usuario']
        }), 200
        
    except Exception as e:
        logger.error(f"Fehler bei der Token-Validierung: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Mitarbeiter Endpunkte
@app.route('/employees', methods=['GET'])
@token_required
def get_employees(current_user):
    """Alle Mitarbeiter abrufen"""
    try:
        employees = db_manager.get_all_employees()
        
        response = jsonify(employees)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Mitarbeiter: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees/with-salaries', methods=['GET'])
@token_required
def get_employees_with_salaries(current_user):
    """Alle Mitarbeiter mit Gehaltsdaten abrufen"""
    try:
        employees = db_manager.get_all_employees_with_salaries()
        
        response = jsonify(employees)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        logger.error(f"Fehler beim Abrufen der Mitarbeiter mit Gehältern: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees/<int:employee_id>', methods=['GET'])
@token_required
def get_employee(current_user, employee_id):
    """Vollständige Mitarbeiterinformationen abrufen"""
    try:
        employee_info = db_manager.get_employee_complete_info(employee_id)
        if not employee_info:
            return jsonify({"error": "Mitarbeiter nicht gefunden"}), 404
        
        response = jsonify(employee_info)
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
        return response
    except Exception as e:
        logger.error(f"Fehler beim Abrufen des Mitarbeiters {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees', methods=['POST'])
@token_required
def create_employee(current_user):
    """Neuen Mitarbeiter erstellen"""
    try:
        employee = request.get_json()
        new_id = db_manager.add_employee(employee)
        if new_id <= 0:
            return jsonify({"error": "Fehler beim Erstellen des Mitarbeiters"}), 400
        
        # Log creation with new details format
        try:
            change_details = db_manager.create_change_details(
                old_data=None,
                new_data=employee
            )
            
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="create",
                objeto="employee",
                id_empleado=new_id,
                details=change_details,
            )
        except Exception as e:
            logger.error(f"Fehler beim Loggen der Mitarbeiter-Erstellung: {e}")
        
        # Mitarbeiter mit neuer ID abrufen
        employees = db_manager.get_all_employees()
        for emp in employees:
            if emp['id_empleado'] == new_id:
                return jsonify(emp)
        
        return jsonify({"error": "Neuer Mitarbeiter nicht gefunden"}), 404
    except Exception as e:
        logger.error(f"Fehler beim Erstellen des Mitarbeiters: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees/<int:employee_id>', methods=['PUT'])
@token_required
def update_employee(current_user, employee_id):
    """Mitarbeiter aktualisieren"""
    try:
        employee = request.get_json()
        # Filter out None values
        employee_data = {k: v for k, v in employee.items() if v is not None}
        if not employee_data:
            return jsonify({"error": "Keine Daten zum Aktualisieren angegeben"}), 400
        
        # Get old data before update
        old_employee = db_manager.get_employee(employee_id)
        if not old_employee:
            return jsonify({"error": "Mitarbeiter nicht gefunden"}), 404
        
        success = db_manager.update_employee(employee_id, 't001_empleados', employee_data)
        if not success:
            return jsonify({"error": "Fehler beim Aktualisieren des Mitarbeiters"}), 400

        # Create change details with old/new values
        change_details = db_manager.create_change_details(
            old_data=old_employee,
            new_data=employee_data
        )

        db_manager.insert_registro_procesamiento(
            usuario_login=current_user,
            accion="update",
            objeto="employee",
            id_empleado=employee_id,
            details=change_details,
        )
        
        return jsonify({"message": "Mitarbeiter erfolgreich aktualisiert"})
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren des Mitarbeiters {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees/<int:employee_id>', methods=['DELETE'])
@token_required
def delete_employee(current_user, employee_id):
    """Mitarbeiter löschen"""
    try:
        success = db_manager.delete_employee(employee_id)
        if not success:
            return jsonify({"error": "Fehler beim Löschen des Mitarbeiters"}), 400
        
        return jsonify({"message": "Mitarbeiter erfolgreich gelöscht"})
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Mitarbeiters {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees/search/<search_term>', methods=['GET'])
@token_required
def search_employees(current_user, search_term):
    """Mitarbeiter suchen"""
    try:
        employees = db_manager.search_employees(search_term)
        return jsonify(employees)
    except Exception as e:
        logger.error(f"Fehler bei der Mitarbeitersuche: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Gehalts Endpunkte
@app.route('/employees/<int:employee_id>/salaries', methods=['POST'])
@token_required
def add_salary(current_user, employee_id):
    """Gehalt für Mitarbeiter hinzufügen"""
    try:
        salary = request.get_json()
        success = db_manager.add_salary(employee_id, salary)
        if not success:
            return jsonify({"error": "Fehler beim Hinzufügen des Gehalts"}), 400

        try:
            change_details = db_manager.create_change_details(
                old_data=None,
                new_data=salary
            )
            
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="create",
                objeto="salary",
                id_empleado=employee_id,
                anio=salary.get('anio'),
                details=change_details,
            )
        except Exception:
            pass
        
        return jsonify({"message": "Gehalt erfolgreich hinzugefügt"})
    except Exception as e:
        logger.error(f"Fehler beim Hinzufügen des Gehalts für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees/<int:employee_id>/salaries/<int:year>', methods=['PUT'])
@token_required
def update_salary(current_user, employee_id, year):
    """Gehalt für Mitarbeiter aktualisieren"""
    try:
        salary = request.get_json()
        
        # Get old salary data before update
        old_salary = db_manager.get_salary(employee_id, year)
        
        # Check if this is actually a create (no existing record)
        is_create = old_salary is None
        
        success = db_manager.update_salary(employee_id, year, salary)
        if not success:
            return jsonify({"error": "Fehler beim Aktualisieren des Gehalts"}), 400

        try:
            # Get the new salary data for comparison
            new_salary = db_manager.get_salary(employee_id, year)
            
            # Create change details
            if is_create:
                # This was actually a create operation
                change_details = db_manager.create_change_details(
                    old_data=None,
                    new_data=salary
                )
                action = "create"
            else:
                # This was an update operation - use the actual database data
                change_details = db_manager.create_change_details(
                    old_data=old_salary,
                    new_data=new_salary
                )
                action = "update"
                
                # Debug: log what we're comparing
                logger.info(f"Salary update comparison - old: {old_salary}, new: {new_salary}, details: {change_details}")

            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion=action,
                objeto="salary",
                id_empleado=employee_id,
                anio=year,
                details=change_details,
            )
        except Exception:
            pass
        
        return jsonify({"message": "Gehalt erfolgreich aktualisiert"})
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren des Gehalts für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/employees/<int:employee_id>/salaries/<int:year>', methods=['DELETE'])
@token_required
def delete_salary(current_user, employee_id, year):
    """Gehalt für Mitarbeiter löschen"""
    try:
        success = db_manager.delete_salary(employee_id, year)
        if not success:
            return jsonify({"error": "Fehler beim Löschen des Gehalts"}), 400
        
        return jsonify({"message": "Gehalt erfolgreich gelöscht"})
    except Exception as e:
        logger.error(f"Fehler beim Löschen des Gehalts für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500



@app.route('/employees/<int:employee_id>/fte', methods=['GET'])
@token_required
def get_employee_fte(current_user, employee_id):
    try:
        rows = db_manager.get_employee_fte(employee_id)
        return jsonify({"items": rows})
    except Exception as e:
        logger.error(f"Fehler beim Abrufen FTE für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500



@app.route('/employees/<int:employee_id>/fte', methods=['PUT'])
@token_required
def upsert_employee_fte(current_user, employee_id):
    try:
        payload = request.get_json() or {}
        year = payload.get('anio')
        month = payload.get('mes')
        porcentaje = payload.get('porcentaje')

        if year is None or month is None or porcentaje is None:
            return jsonify({"error": "anio, mes und porcentaje sind erforderlich"}), 400

        old_rows = db_manager.get_employee_fte(employee_id)
        old_value = None
        for r in old_rows:
            if int(r.get('anio')) == int(year) and int(r.get('mes')) == int(month):
                old_value = r
                break

        success = db_manager.upsert_employee_fte(employee_id, int(year), int(month), float(porcentaje))
        if not success:
            return jsonify({"error": "Fehler beim Speichern der Stundenreduzierung"}), 400

        try:
            new_value = {"anio": int(year), "mes": int(month), "porcentaje": float(porcentaje)}
            change_details = db_manager.create_change_details(old_data=old_value, new_data=new_value)
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="update" if old_value else "create",
                objeto="fte",
                id_empleado=employee_id,
                anio=int(year),
                mes=int(month),
                details=change_details,
            )
        except Exception:
            pass

        return jsonify({"message": "Stundenreduzierung gespeichert"})
    except Exception as e:
        logger.error(f"Fehler beim Speichern FTE für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500



@app.route('/employees/<int:employee_id>/fte/<int:year>/<int:month>', methods=['DELETE'])
@token_required
def delete_employee_fte(current_user, employee_id, year, month):
    try:
        old_rows = db_manager.get_employee_fte(employee_id)
        old_value = None
        for r in old_rows:
            if int(r.get('anio')) == int(year) and int(r.get('mes')) == int(month):
                old_value = r
                break

        success = db_manager.delete_employee_fte(employee_id, year, month)
        if not success:
            return jsonify({"error": "Fehler beim Löschen der Stundenreduzierung"}), 400

        try:
            change_details = db_manager.create_change_details(old_data=old_value, new_data=None)
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="delete",
                objeto="fte",
                id_empleado=employee_id,
                anio=int(year),
                mes=int(month),
                details=change_details,
            )
        except Exception:
            pass

        return jsonify({"message": "Stundenreduzierung gelöscht"})
    except Exception as e:
        logger.error(f"Fehler beim Löschen FTE für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Einkünfte Endpunkte
@app.route('/employees/<int:employee_id>/ingresos/<int:year>', methods=['PUT'])
@token_required
def update_ingresos(current_user, employee_id, year):
    """Bruttoeinkünfte für Mitarbeiter aktualisieren"""
    try:
        ingresos = request.get_json()
        success = db_manager.update_ingresos(employee_id, year, ingresos)
        if not success:
            return jsonify({"error": "Fehler beim Aktualisieren der Bruttoeinkünfte"}), 400

        try:
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="update",
                objeto="ingresos",
                id_empleado=employee_id,
                anio=year,
                details=ingresos,
            )
        except Exception:
            pass
        
        return jsonify({"message": "Bruttoeinkünfte erfolgreich aktualisiert"})
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren der Bruttoeinkünfte für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Abzüge Endpunkte
@app.route('/employees/<int:employee_id>/deducciones/<int:year>', methods=['PUT'])
@token_required
def update_deducciones(current_user, employee_id, year):
    """Abzüge für Mitarbeiter aktualisieren"""
    try:
        deducciones = request.get_json()
        success = db_manager.update_deducciones(employee_id, year, deducciones)
        if not success:
            return jsonify({"error": "Fehler beim Aktualisieren der Abzüge"}), 400

        try:
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="update",
                objeto="deducciones",
                id_empleado=employee_id,
                anio=year,
                details=deducciones,
            )
        except Exception:
            pass
        
        return jsonify({"message": "Abzüge erfolgreich aktualisiert"})
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren der Abzüge für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Excel Export Endpunkt
@app.route('/export/excel/<int:year>', methods=['GET'])
@app.route('/export/excel/<int:year>/<int:month>', methods=['GET'])
@token_required
def export_excel(current_user, year, month=None):
    """Excel-Export für Gehaltsdaten - jährlich oder monatlich"""
    try:
        extra_raw = request.args.get('extra', '').strip().lower()
        extra = extra_raw in {'1', 'true', 'yes', 'on'}

        # Temporäre Datei erstellen
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        if month:
            download_filename = f"NOMINA_TOTAL_EXTRA_{year}_{month}.xlsx" if extra else f"NOMINA_TOTAL_{year}_{month}.xlsx"
            temp_filename = f"nomina_total_{year}_{month}_{timestamp}.xlsx"
        else:
            download_filename = f"gehaltsabrechnung_{year}.xlsx"
            temp_filename = f"gehaltsabrechnung_{year}_{timestamp}.xlsx"
        filepath = f"C:/temp/{temp_filename}"
        
        # Temp-Verzeichnis erstellen falls nicht vorhanden
        os.makedirs("C:/temp", exist_ok=True)
        
        success = db_manager.export_nomina_excel(year, filepath, month, extra=extra)
        if not success:
            return jsonify({"error": "Fehler beim Excel-Export"}), 400
        
        # Datei zurückgeben
        return send_file(
            filepath,
            as_attachment=True,
            download_name=download_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Fehler beim Excel-Export für Jahr {year}, Monat {month}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Asiento Nomina Export Endpunkt
@app.route('/export/asiento_nomina/<int:year>/<int:month>', methods=['GET'])
@token_required
def export_asiento_nomina(current_user, year, month):
    """Asiento Nomina Excel-Export für Gehaltsdaten - monatlich"""
    try:
        # Temporäre Datei erstellen
        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        download_filename = f"ASIENTO_NOMINA_{year}_{month}.xlsx"
        temp_filename = f"asiento_nomina_{year}_{month}_{timestamp}.xlsx"
        filepath = f"C:/temp/{temp_filename}"
        
        # Temp-Verzeichnis erstellen falls nicht vorhanden
        os.makedirs("C:/temp", exist_ok=True)
        
        success = db_manager.export_asiento_nomina_excel(year, month, filepath)
        if not success:
            return jsonify({"error": "Fehler beim Asiento Nomina Export"}), 400
        
        # Datei zurückgeben
        return send_file(
            filepath,
            as_attachment=True,
            download_name=download_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Fehler beim Asiento Nomina Export für Jahr {year}, Monat {month}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500


# IRPF Export Endpunkt
@app.route('/export/irpf/<int:year>', methods=['GET'])
@app.route('/export/irpf/<int:year>/<int:month>', methods=['GET'])
@token_required
def export_irpf(current_user, year, month=None):
    try:
        extra_raw = request.args.get('extra', '').strip().lower()
        extra = extra_raw in {'1', 'true', 'yes', 'on'}

        timestamp = datetime.now(timezone.utc).strftime("%Y%m%d_%H%M%S")
        if month:
            download_filename = f"IRPF_EXTRA_{year}_{month}.xlsx" if extra else f"IRPF_{year}_{month}.xlsx"
            temp_filename = f"irpf_{year}_{month}_{timestamp}.xlsx"
        else:
            download_filename = f"IRPF_{year}.xlsx"
            temp_filename = f"irpf_{year}_{timestamp}.xlsx"
        filepath = f"C:/temp/{temp_filename}"

        os.makedirs("C:/temp", exist_ok=True)

        success = db_manager.export_irpf_excel(year=year, output_path=filepath, month=month, extra=extra)
        if not success:
            return jsonify({"error": "Fehler beim IRPF Export"}), 400

        return send_file(
            filepath,
            as_attachment=True,
            download_name=download_filename,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
    except Exception as e:
        logger.error(f"Fehler beim IRPF Export für Jahr {year}, Monat {month}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500



# Carry Over Endpunkte
@app.route('/carry-over/<int:employee_id>/<int:year>/<int:month>', methods=['GET'])
@token_required
def list_carry_over(current_user, employee_id, year, month):
    try:
        items = db_manager.list_carry_over_by_source(employee_id, year, month)
        return jsonify({"items": items})
    except Exception as e:
        logger.error(f"Fehler beim Abrufen Carry Over für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500


@app.route('/carry-over', methods=['POST'])
@token_required
def create_carry_over(current_user):
    try:
        payload = request.get_json() or {}
        employee_id = payload.get('employee_id')
        year = payload.get('year')
        month = payload.get('month')
        items = payload.get('items') or []
        defer_concepts = payload.get('defer_concepts') or []

        if employee_id is None or year is None or month is None:
            return jsonify({"error": "employee_id, year und month sind erforderlich"}), 400

        try:
            employee_id_i = int(employee_id)
            year_i = int(year)
            month_i = int(month)
        except Exception:
            return jsonify({"error": "employee_id, year und month müssen Ganzzahlen sein"}), 400

        if month_i < 1 or month_i > 12:
            return jsonify({"error": "month muss zwischen 1 und 12 liegen"}), 400

        if not isinstance(items, list):
            return jsonify({"error": "items muss eine Liste sein"}), 400
        if defer_concepts is not None and not isinstance(defer_concepts, list):
            return jsonify({"error": "defer_concepts muss eine Liste sein"}), 400

        success = db_manager.create_carry_over_batch(
            employee_id=employee_id_i,
            source_year=year_i,
            source_month=month_i,
            items=items,
            defer_concepts=defer_concepts,
        )

        if not success:
            return jsonify({"error": "Fehler beim Speichern Carry Over"}), 400

        try:
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="create",
                objeto="carry_over",
                id_empleado=employee_id_i,
                anio=year_i,
                mes=month_i,
                details={"items": items},
            )
        except Exception:
            pass

        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Fehler beim Speichern Carry Over: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500


@app.route('/carry-over/<int:carry_over_id>', methods=['DELETE'])
@token_required
def delete_carry_over(current_user, carry_over_id):
    try:
        success = db_manager.delete_carry_over(carry_over_id)
        if not success:
            return jsonify({"error": "Fehler beim Löschen Carry Over"}), 400
        return jsonify({"success": True})
    except Exception as e:
        logger.error(f"Fehler beim Löschen Carry Over {carry_over_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500



# Import Endpunkte
@app.route('/imports/horas-dietas', methods=['POST'])
@token_required
def import_horas_dietas(current_user):
    try:
        year_raw = request.form.get('year', '').strip()
        month_raw = request.form.get('month', '').strip()
        if not year_raw or not month_raw:
            return jsonify({"error": "year und month sind erforderlich"}), 400

        try:
            year = int(year_raw)
            month = int(month_raw)
        except Exception:
            return jsonify({"error": "year und month müssen Ganzzahlen sein"}), 400

        if month < 1 or month > 12:
            return jsonify({"error": "month muss zwischen 1 und 12 liegen"}), 400

        if 'file' not in request.files:
            return jsonify({"error": "file ist erforderlich"}), 400
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({"error": "Ungültige Datei"}), 400

        try:
            content = file.read()
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"Fehler beim Lesen der Excel-Datei: {e}")
            return jsonify({"error": "Excel-Datei konnte nicht gelesen werden"}), 400

        result = db_manager.import_horas_dietas_worksheet(
            worksheet=ws,
            year=year,
            month=month,
            usuario_login=current_user,
            source_filename=file.filename,
        )

        status_code = 200 if result.get('success') else 400
        return jsonify(result), status_code

    except Exception as e:
        logger.error(f"Fehler beim Import Horas+Dietas: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500


@app.route('/imports/gasolina', methods=['POST'])
@token_required
def import_gasolina(current_user):
    try:
        year_raw = request.form.get('year', '').strip()
        month_raw = request.form.get('month', '').strip()
        logger.info(f"Gasolina Import received: year='{year_raw}', month='{month_raw}', files={list(request.files.keys())}")
        if not year_raw or not month_raw:
            return jsonify({"error": "year und month sind erforderlich"}), 400

        try:
            year = int(year_raw)
            month = int(month_raw)
        except Exception:
            return jsonify({"error": "year und month müssen Ganzzahlen sein"}), 400

        if month < 1 or month > 12:
            return jsonify({"error": "month muss zwischen 1 und 12 liegen"}), 400

        if 'file' not in request.files:
            return jsonify({"error": "file ist erforderlich"}), 400
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({"error": "Ungültige Datei"}), 400

        try:
            content = file.read()
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"Fehler beim Lesen der Excel-Datei: {e}")
            return jsonify({"error": "Excel-Datei konnte nicht gelesen werden"}), 400

        result = db_manager.import_gasolina_worksheet(
            worksheet=ws,
            year=year,
            month=month,
            usuario_login=current_user,
            source_filename=file.filename,
        )

        status_code = 200 if result.get('success') else 400
        return jsonify(result), status_code

    except Exception as e:
        logger.error(f"Fehler beim Import Gasolina: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500


@app.route('/imports/cotizacion-especie', methods=['POST'])
@token_required
def import_cotizacion_especie(current_user):
    try:
        year_raw = request.form.get('year', '').strip()
        month_raw = request.form.get('month', '').strip()
        if not year_raw or not month_raw:
            return jsonify({"error": "year und month sind erforderlich"}), 400

        try:
            year = int(year_raw)
            month = int(month_raw)
        except Exception:
            return jsonify({"error": "year und month müssen Ganzzahlen sein"}), 400

        if month < 1 or month > 12:
            return jsonify({"error": "month muss zwischen 1 und 12 liegen"}), 400

        if 'file' not in request.files:
            return jsonify({"error": "file ist erforderlich"}), 400
        file = request.files['file']
        if not file or not file.filename:
            return jsonify({"error": "Ungültige Datei"}), 400

        try:
            content = file.read()
            wb = load_workbook(filename=io.BytesIO(content), data_only=True)
            ws = wb.active
        except Exception as e:
            logger.error(f"Fehler beim Lesen der Excel-Datei: {e}")
            return jsonify({"error": "Excel-Datei konnte nicht gelesen werden"}), 400

        result = db_manager.import_cotizacion_especie_worksheet(
            worksheet=ws,
            year=year,
            month=month,
            usuario_login=current_user,
            source_filename=file.filename,
        )

        status_code = 200 if result.get('success') else 400
        return jsonify(result), status_code

    except Exception as e:
        logger.error(f"Fehler beim Import Cotizacion Especie: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Monatliche Einkünfte Endpunkte
@app.route('/employees/<int:employee_id>/ingresos/<int:year>/<int:month>', methods=['PUT'])
@token_required
def update_ingresos_mensuales(current_user, employee_id, year, month):
    """Monatliche Bruttoeinkünfte für Mitarbeiter aktualisieren"""
    try:
        ingresos = request.get_json()
        success = db_manager.update_ingresos_mensuales(employee_id, year, month, ingresos)
        if not success:
            return jsonify({"error": "Fehler beim Aktualisieren der monatlichen Bruttoeinkünfte"}), 400

        try:
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="update",
                objeto="ingresos_mensuales",
                id_empleado=employee_id,
                anio=year,
                mes=month,
                details=ingresos,
            )
        except Exception:
            pass
        
        return jsonify({"message": "Monatliche Bruttoeinkünfte erfolgreich aktualisiert"})
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren der monatlichen Bruttoeinkünfte für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Monatliche Abzüge Endpunkte
@app.route('/employees/<int:employee_id>/deducciones/<int:year>/<int:month>', methods=['PUT'])
@token_required
def update_deducciones_mensuales(current_user, employee_id, year, month):
    """Monatliche Abzüge für Mitarbeiter aktualisieren"""
    try:
        deducciones = request.get_json()
        result = db_manager.update_deducciones_mensuales(employee_id, year, month, deducciones)
        
        if not result["success"]:
            return jsonify({"error": result.get("error", "Fehler beim Aktualisieren der monatlichen Abzüge")}), 400

        try:
            db_manager.insert_registro_procesamiento(
                usuario_login=current_user,
                accion="update",
                objeto="deducciones_mensuales",
                id_empleado=employee_id,
                anio=year,
                mes=month,
                details=deducciones,
            )
        except Exception:
            pass
        
        # Bereite die Antwort mit Propagation-Informationen vor
        response_message = "Monatliche Abzüge erfolgreich aktualisiert"
        if result.get("propagation_info") and result["propagation_info"].get("propagated"):
            months = result["propagation_info"]["months_updated"]
            response_message += f". Cotizacion Especie wurde automatisch für die folgenden Monate übernommen: {', '.join(map(str, months))}"
        
        return jsonify({
            "message": response_message,
            "propagation_info": result.get("propagation_info")
        })
    except Exception as e:
        logger.error(f"Fehler beim Aktualisieren der monatlichen Abzüge für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500



@app.route('/employees/<int:employee_id>/registro_procesamiento', methods=['GET'])
@token_required
def get_registro_procesamiento(current_user, employee_id):
    """Registro de procesamiento für einen Mitarbeiter abrufen"""
    try:
        anio = request.args.get('anio', default=None, type=int)
        mes = request.args.get('mes', default=None, type=int)
        limit = request.args.get('limit', default=200, type=int)

        rows = db_manager.get_registro_procesamiento(employee_id, anio=anio, mes=mes, limit=limit)
        return jsonify({"items": rows})

    except Exception as e:
        logger.error(f"Fehler beim Abrufen registro_procesamiento für Mitarbeiter {employee_id}: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

@app.route('/registro_procesamiento', methods=['GET'])
@token_required
def get_global_registro_procesamiento(current_user):
    """Globale Registro de procesamiento abrufen"""
    try:
        id_empleado = request.args.get('id_empleado', default=None, type=int)
        anio = request.args.get('anio', default=None, type=int)
        mes = request.args.get('mes', default=None, type=int)
        limit = request.args.get('limit', default=200, type=int)

        rows = db_manager.get_global_registro_procesamiento(
            id_empleado=id_empleado, 
            anio=anio, 
            mes=mes, 
            limit=limit
        )
        return jsonify({"items": rows})

    except Exception as e:
        logger.error(f"Fehler beim Abrufen global registro_procesamiento: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Health Check
@app.route('/health', methods=['GET'])
def health_check():
    """Health Check Endpunkt"""
    return jsonify({
        "status": "healthy", 
        "timestamp": datetime.now(timezone.utc).isoformat()
    })



# Settings Endpunkte
@app.route('/settings/payout-month', methods=['GET'])
@token_required
def get_payout_month(current_user):
    """Gibt den globalen Auszahlungsmonat (1-12) zurück"""
    try:
        payout_month = db_manager.get_payout_month()
        return jsonify({"payout_month": payout_month})
    except Exception as e:
        logger.error(f"Fehler beim Lesen von payout_month: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500



@app.route('/settings/payout-month', methods=['PUT'])
@token_required
def set_payout_month(current_user):
    """Setzt den globalen Auszahlungsmonat (1-12)"""
    try:
        data = request.get_json() or {}
        payout_month = data.get('payout_month')

        if payout_month is None or not isinstance(payout_month, int) or payout_month < 1 or payout_month > 12:
            return jsonify({"error": "payout_month muss eine Zahl von 1 bis 12 sein"}), 400

        success = db_manager.set_payout_month(payout_month)
        if not success:
            return jsonify({"error": "Konnte payout_month nicht speichern"}), 400

        return jsonify({"success": True, "payout_month": payout_month})

    except Exception as e:
        logger.error(f"Fehler beim Setzen von payout_month: {e}")
        return jsonify({"error": "Interner Serverfehler"}), 500

# Atrasos Neuberechnung Endpunkt
@app.route('/settings/recalculate-atrasos', methods=['POST'])
@token_required
def recalculate_atrasos(current_user):
    """Berechnet alle Atrasos für ein bestimmtes Jahr neu"""
    try:
        data = request.get_json() or {}
        year = data.get('year')
        
        if year is None or not isinstance(year, int) or year < 2020 or year > 2030:
            return jsonify({"error": "year muss eine gültige Jahreszahl zwischen 2020 und 2030 sein"}), 400
        
        # Führe die Neuberechnung durch
        result = db_manager.recalculate_all_atrasos_for_year(year)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify(result), 400
            
    except Exception as e:
        logger.error(f"Fehler bei der Neuberechnung der Atrasos: {e}")
        return jsonify({
            "success": False,
            "message": "Interner Serverfehler bei der Neuberechnung",
            "updated_count": 0,
            "total_count": 0,
            "errors": [str(e)]
        }), 500



# Bulk Ingresos/Deducciones für alle aktiven Mitarbeiter
@app.route('/settings/apply-ingresos-deducciones', methods=['POST'])
@token_required
def apply_ingresos_deducciones_to_all_active(current_user):
    """Setzt Zulagen (ingresos) und/oder Abzüge (deducciones) für ein Jahr für alle aktiven Mitarbeiter"""
    try:
        data = request.get_json() or {}
        year = data.get('year')
        ingresos = data.get('ingresos')
        deducciones = data.get('deducciones')
        categoria = data.get('categoria')

        if year is None or not isinstance(year, int) or year < 2020 or year > 2030:
            return jsonify({"error": "year muss eine gültige Jahreszahl zwischen 2020 und 2030 sein"}), 400

        if ingresos is None and deducciones is None:
            return jsonify({"error": "Mindestens ingresos oder deducciones muss gesetzt sein"}), 400

        result = db_manager.apply_yearly_ingresos_and_deducciones_to_all_active(
            year=year,
            ingresos=ingresos,
            deducciones=deducciones,
            categoria=categoria,
        )

        if result.get('success'):
            return jsonify(result), 200
        return jsonify(result), 400

    except Exception as e:
        logger.error(f"Fehler beim Bulk-Setzen von ingresos/deducciones: {e}")
        return jsonify({
            "success": False,
            "message": "Interner Serverfehler",
            "updated_count": 0,
            "total_count": 0,
            "errors": [str(e)]
        }), 500

# Gehaltskopierung Endpunkte
@app.route('/salaries/copy-to-year/<int:target_year>', methods=['POST'])
@token_required
def copy_salaries_to_year(current_user, target_year):
    """Kopiert Gehälter vom Vorjahr ins Zieljahr"""
    try:
        result = db_manager.copy_salaries_to_new_year(target_year)
        return jsonify(result)
    except Exception as e:
        logger.error(f"Fehler bei der Gehaltskopierung für Jahr {target_year}: {e}")
        return jsonify({
            "success": False,
            "message": "Interner Serverfehler bei der Gehaltskopierung",
            "copied_count": 0,
            "skipped_count": 0,
            "errors": [str(e)]
        }), 500

@app.route('/salaries/missing-years', methods=['GET'])
@token_required
def get_missing_salary_years(current_user):
    """Gibt Jahre zurück, für die Mitarbeiter keine Gehälter haben"""
    try:
        missing_years = db_manager.get_missing_salary_years()
        return jsonify({
            "success": True,
            "missing_years": missing_years
        })
    except Exception as e:
        logger.error(f"Fehler bei der Abfrage fehlender Gehaltsjahre: {e}")
        return jsonify({
            "success": False,
            "message": "Interner Serverfehler bei der Abfrage",
            "missing_years": []
        }), 500

# Gehaltserhöhung Endpunkt
@app.route('/salaries/percentage-increase', methods=['POST'])
@token_required
def apply_percentage_increase(current_user):
    """Wendet eine prozentuale Gehaltserhöhung auf alle aktiven Mitarbeiter an"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "Keine Daten im Request Body gefunden"}), 400
        
        target_year = data.get('target_year')
        percentage_increase = data.get('percentage_increase')
        absolute_increase = data.get('absolute_increase')
        excluded_employee_ids = data.get('excluded_employee_ids', [])
        
        if target_year is None:
            return jsonify({"error": "target_year ist erforderlich"}), 400
        
        if percentage_increase is None and absolute_increase is None:
            return jsonify({"error": "Entweder percentage_increase oder absolute_increase ist erforderlich"}), 400
        
        if not isinstance(target_year, int) or target_year < 2020:
            return jsonify({"error": "target_year muss eine gültige Jahreszahl sein"}), 400
        
        if percentage_increase is not None and (not isinstance(percentage_increase, (int, float)) or percentage_increase <= 0):
            return jsonify({"error": "percentage_increase muss eine positive Zahl sein"}), 400
        
        if absolute_increase is not None and (not isinstance(absolute_increase, (int, float)) or absolute_increase <= 0):
            return jsonify({"error": "absolute_increase muss eine positive Zahl sein"}), 400
        
        if not isinstance(excluded_employee_ids, list):
            return jsonify({"error": "excluded_employee_ids muss eine Liste sein"}), 400
        
        # Wende die Gehaltserhöhung an
        result = db_manager.apply_percentage_salary_increase(target_year, percentage_increase, absolute_increase, excluded_employee_ids)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify(result), 400
            
    except Exception as e:
        logger.error(f"Fehler bei prozentualer Gehaltserhöhung: {e}")
        return jsonify({
            "success": False,
            "message": "Interner Serverfehler",
            "errors": [str(e)]
        }), 500

# Gehaltserhöhung für einzelnen Mitarbeiter Endpunkt
@app.route('/employees/<int:employee_id>/salary-increase', methods=['POST'])
@token_required
def apply_employee_salary_increase(current_user, employee_id):
    """Wendet eine prozentuale Gehaltserhöhung auf einen einzelnen Mitarbeiter an"""
    try:
        data = request.get_json()
        
        if not data:
            return jsonify({"error": "Keine Daten im Request Body gefunden"}), 400
        
        target_year = data.get('target_year')
        percentage_increase = data.get('percentage_increase')
        
        if target_year is None or percentage_increase is None:
            return jsonify({"error": "target_year und percentage_increase sind erforderlich"}), 400
        
        if not isinstance(target_year, int) or target_year < 2020:
            return jsonify({"error": "target_year muss eine gültige Jahreszahl sein"}), 400
        
        if not isinstance(percentage_increase, (int, float)) or percentage_increase <= 0:
            return jsonify({"error": "percentage_increase muss eine positive Zahl sein"}), 400
        
        # Wende die Gehaltserhöhung für einen Mitarbeiter an
        result = db_manager.apply_employee_salary_increase(employee_id, target_year, percentage_increase)
        
        if result['success']:
            return jsonify(result), 200
        else:
            return jsonify(result), 400
            
    except Exception as e:
        logger.error(f"Fehler bei prozentualer Gehaltserhöhung für Mitarbeiter {employee_id}: {e}")
        return jsonify({
            "success": False,
            "message": "Interner Serverfehler",
            "errors": [str(e)]
        }), 500

if __name__ == "__main__":
    print("Starte Flask API Server auf http://localhost:8000")
    print("API-Dokumentation: http://localhost:8000/health")
    print("Health Check: http://localhost:8000/health")
    
    app.run(host="0.0.0.0", port=8000, debug=True)
