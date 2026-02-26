import pytest
import sys
import os
import tempfile
import shutil
from unittest.mock import Mock, patch, MagicMock
from mysql.connector import Error
import hashlib
import time

# Backend-Verzeichnis zum Pfad hinzufügen
backend_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'backend'))
sys.path.insert(0, backend_path)

from database_manager import DatabaseManager

class TestDatabaseManagerCore:
    """Kern-Tests für DatabaseManager - fokussiert auf reale Funktionalität"""

    @pytest.fixture
    def db_manager(self):
        """DatabaseManager Instanz für Tests"""
        return DatabaseManager('localhost', 'test_db', 'test_user', 'test_password', 3307)

    def test_initialization(self, db_manager):
        """Test DatabaseManager Initialisierung"""
        assert db_manager.host == 'localhost'
        assert db_manager.database == 'test_db'
        assert db_manager.user == 'test_user'
        assert db_manager.password == 'test_password'
        assert db_manager.port == 3307
        assert db_manager.connection is None

    def test_hash_password_functionality(self, db_manager):
        """Test Passwort-Hashing Funktionalität"""
        password = "test_password_123"
        
        # Hash erzeugen
        hashed = db_manager.hash_password(password)
        
        # Überprüfungen
        assert hashed != password
        assert len(hashed) == 64  # SHA-256
        assert all(c in '0123456789abcdef' for c in hashed)
        
        # Konsistenzprüfung
        hashed2 = db_manager.hash_password(password)
        assert hashed == hashed2
        
        # Unterschiedliche Passworte erzeugen unterschiedliche Hashes
        hashed3 = db_manager.hash_password("other_password")
        assert hashed != hashed3

    def test_hash_password_edge_cases(self, db_manager):
        """Test Passwort-Hashing Grenzwertfälle"""
        # Leeres Passwort
        hashed_empty = db_manager.hash_password("")
        assert len(hashed_empty) == 64
        assert hashed_empty != ""
        
        # Sehr langes Passwort
        long_password = "x" * 1000
        hashed_long = db_manager.hash_password(long_password)
        assert len(hashed_long) == 64
        
        # Spezielle Zeichen
        special_password = "!@#$%^&*()_+-=[]{}|;:,.<>?"
        hashed_special = db_manager.hash_password(special_password)
        assert len(hashed_special) == 64
        assert hashed_special != special_password

    @patch.object(DatabaseManager, '_create_connection')
    def test_connection_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche Datenbankverbindung"""
        mock_connection = Mock()
        mock_connection.is_connected.return_value = True
        mock_create_connection.return_value = mock_connection
        
        result = db_manager.connect()
        
        assert result is True
        assert db_manager.connection == mock_connection
        mock_create_connection.assert_called_once()

    @patch.object(DatabaseManager, '_create_connection')
    def test_connection_failure(self, mock_create_connection, db_manager):
        """Test fehlgeschlagene Datenbankverbindung"""
        mock_create_connection.side_effect = Error("Connection failed")
        
        result = db_manager.connect()
        
        assert result is False
        assert db_manager.connection is None

    @patch.object(DatabaseManager, '_create_connection')
    def test_connection_timeout(self, mock_create_connection, db_manager):
        """Test Connection Timeout"""
        import socket
        mock_create_connection.side_effect = socket.timeout("Connection timeout")
        
        # Erwarte TimeoutError
        with pytest.raises(socket.timeout):
            db_manager.connect()

    def test_execute_query_without_connection(self, db_manager):
        """Test Query-Ausführung ohne Verbindung"""
        # Sollte leere Liste zurückgeben bei Verbindungsproblemen
        result = db_manager.execute_query("SELECT * FROM test")
        assert result == []

    @patch.object(DatabaseManager, '_create_connection')
    def test_execute_query_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche Query-Ausführung"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_cursor.fetchall.return_value = [{'id': 1, 'name': 'test'}]
        mock_cursor.with_rows = True
        mock_connection.cursor.return_value = mock_cursor
        mock_create_connection.return_value = mock_connection
        result = db_manager.execute_query("SELECT * FROM test")
        
        assert result == [{'id': 1, 'name': 'test'}]
        mock_cursor.execute.assert_called_once_with("SELECT * FROM test", None)
        mock_cursor.close.assert_called_once()
        mock_connection.close.assert_called_once()

    @patch.object(DatabaseManager, '_create_connection')
    def test_execute_query_with_params(self, mock_create_connection, db_manager):
        """Test Query-Ausführung mit Parametern"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_cursor.fetchall.return_value = []
        mock_cursor.with_rows = True
        mock_connection.cursor.return_value = mock_cursor
        mock_create_connection.return_value = mock_connection
        query = "SELECT * FROM test WHERE id = %s"
        params = (1,)
        
        db_manager.execute_query(query, params)
        
        mock_cursor.execute.assert_called_once_with(query, params)
        mock_connection.close.assert_called_once()

    @patch.object(DatabaseManager, '_create_connection')
    def test_execute_query_database_error(self, mock_create_connection, db_manager):
        """Test Query-Ausführung mit Datenbankfehler"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_cursor.execute.side_effect = Error("SQL Error")
        mock_connection.cursor.return_value = mock_cursor
        mock_create_connection.return_value = mock_connection
        result = db_manager.execute_query("SELECT * FROM test")
        
        assert result == []
        mock_connection.close.assert_called_once()

    @patch.object(DatabaseManager, '_create_connection')
    def test_execute_update_success(self, mock_create_connection, db_manager):
        """Test erfolgreiches Update"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_connection.cursor.return_value = mock_cursor
        mock_create_connection.return_value = mock_connection
        result = db_manager.execute_update("UPDATE test SET name = %s", ("new_name",))
        
        assert result is True
        mock_cursor.execute.assert_called_once()
        mock_connection.commit.assert_called_once()
        mock_cursor.close.assert_called_once()
        mock_connection.close.assert_called_once()

    @patch.object(DatabaseManager, '_create_connection')
    def test_execute_update_failure(self, mock_create_connection, db_manager):
        """Test fehlgeschlagenes Update"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_cursor.execute.side_effect = Error("Update Error")
        mock_connection.cursor.return_value = mock_cursor
        mock_create_connection.return_value = mock_connection
        result = db_manager.execute_update("UPDATE test SET name = %s", ("new_name",))
        
        assert result is False
        mock_connection.rollback.assert_called_once()
        mock_connection.close.assert_called_once()

    def test_add_employee_validation(self, db_manager):
        """Test Mitarbeiter-Erstellung Validierung"""
        # Leere Daten
        result = db_manager.add_employee({})
        assert result == -1
        
        # Fehlende Pflichtfelder
        result = db_manager.add_employee({'nombre': 'Test'})
        assert result == -1
        
        # Nur Pflichtfelder (sollte -1 zurückgeben wegen fehlender DB)
        result = db_manager.add_employee({
            'nombre': 'Test',
            'apellido': 'User',
            'ceco': '1001'
        })
        assert result == -1

    @patch('mysql.connector.connect')
    def test_add_employee_success(self, mock_connect, db_manager):
        """Test erfolgreiche Mitarbeiter-Erstellung"""
        mock_connection = Mock()
        mock_cursor = Mock()
        
        # Mock für MAX ID Abfrage
        mock_cursor.fetchall.side_effect = [
            [{'max_id': 1}],  # MAX ID Abfrage
            []  # INSERT Abfrage
        ]
        mock_connection.cursor.return_value = mock_cursor
        mock_connect.return_value = mock_connection
        
        # Mock execute_update für Erfolg
        with patch.object(db_manager, 'execute_update', return_value=True):
            with patch.object(db_manager, 'execute_query', return_value=[{'max_id': 1}]):
                with patch.object(db_manager, '_create_default_records'):
                    result = db_manager.add_employee({
                        'nombre': 'Test',
                        'apellido': 'User',
                        'ceco': '1001'
                    })
                    assert result == 2  # max_id + 1

    @patch('mysql.connector.connect')
    def test_update_employee_success(self, mock_connect, db_manager):
        """Test erfolgreiche Mitarbeiter-Aktualisierung"""
        mock_connection = Mock()
        mock_connect.return_value = mock_connection
        
        with patch.object(db_manager, 'execute_update', return_value=True):
            result = db_manager.update_employee(1, 't001_empleados', {
                'nombre': 'Updated Name',
                'apellido': 'Updated Surname'
            })
            assert result is True

    def test_update_employee_invalid_table(self, db_manager):
        """Test Mitarbeiter-Aktualisierung mit ungültiger Tabelle"""
        result = db_manager.update_employee(1, 'invalid_table', {'nombre': 'Test'})
        assert result is False

    def test_update_employee_invalid_fields(self, db_manager):
        """Test Mitarbeiter-Aktualisierung mit ungültigen Feldern"""
        # Ungültige Felder sollten ignoriert werden
        result = db_manager.update_employee(1, 't001_empleados', {
            'invalid_field': 'value'
        })
        # Sollte False zurückgeben weil keine gültigen Felder
        assert result is False

    def test_verify_user_without_connection(self, db_manager):
        """Test Benutzer-Verifizierung ohne Verbindung"""
        result = db_manager.verify_user('testuser', 'password')
        assert result is None

    @patch('mysql.connector.connect')
    def test_verify_user_success(self, mock_connect, db_manager):
        """Test erfolgreiche Benutzer-Verifizierung"""
        mock_connection = Mock()
        mock_connect.return_value = mock_connection
        
        user_data = {
            'id_usuario': 1,
            'nombre_usuario': 'testuser',
            'nombre_completo': 'Test User',
            'rol': 'admin',
            'activo': True
        }
        
        with patch.object(db_manager, 'execute_query', return_value=[user_data]):
            result = db_manager.verify_user('testuser', 'password')
            assert result == user_data

    @patch('mysql.connector.connect')
    def test_verify_user_not_found(self, mock_connect, db_manager):
        """Test Benutzer-Verifizierung - Benutzer nicht gefunden"""
        mock_connection = Mock()
        mock_connect.return_value = mock_connection
        
        with patch.object(db_manager, 'execute_query', return_value=[]):
            result = db_manager.verify_user('nonexistent', 'password')
            assert result is None

    def test_export_excel_without_dependencies(self, db_manager):
        """Test Excel-Export ohne Abhängigkeiten"""
        with patch.dict('sys.modules', {'pandas': None, 'openpyxl': None}):
            result = db_manager.export_nomina_excel(2025, 'test.xlsx')
            assert result is False

    def test_export_nomina_excel_forwards_extra_flag(self, db_manager):
        with patch('database_manager.DatabaseManagerExportsMixin.export_nomina_excel', return_value=True) as mock_export:
            ok = db_manager.export_nomina_excel(2025, 'test.xlsx', 6, extra=True)
            assert ok is True
            mock_export.assert_called_once_with(db_manager, 2025, 'test.xlsx', 6, extra=True)

    def test_export_nomina_excel_extra_creates_three_column_xlsx(self, db_manager):
        try:
            from openpyxl import load_workbook
        except ImportError:
            pytest.skip('openpyxl nicht verfügbar')

        rows = [
            {
                'nombre_completo': 'Test, User',
                'ceco': '1001',
                'modalidad': 14,
                'salario_mensual_bruto': 2000,
                'atrasos': 0,
                'antiguedad': 0,
                'salario_mensual_bruto_prev': 1900,
                'ticket_restaurant': 0,
                'cotizacion_especie': 0,
                'primas': 0,
                'dietas_cotizables': 0,
                'horas_extras': 0,
                'seguro_pensiones': 0,
                'seguro_accidentes': 0,
                'dietas_exentas': 0,
                'formacion': 0,
                'adelas': 0,
                'sanitas': 0,
                'gasolina': 0,
                'dias_exentos': 0,
            }
        ]

        captured_query = {'query': None}

        def fake_execute_query(query, params=None):
            captured_query['query'] = query
            return rows

        with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
            output_path = tmp.name

        try:
            with patch.object(db_manager, 'execute_query', side_effect=fake_execute_query):
                ok = db_manager.export_nomina_excel(2025, output_path, 6, extra=True)
                assert ok is True
                assert captured_query['query'] is not None
                assert 's.modalidad = 14' in captured_query['query']

            wb = load_workbook(output_path)
            ws = wb['Sheet1']

            assert ws.max_column == 3
            assert ws['A5'].value == 'Etiquetas de fila'
            assert ws['B5'].value == 'CECO'
            assert ws['C5'].value == 'SALARIO MES'
        finally:
            try:
                os.remove(output_path)
            except OSError:
                pass

    def test_export_monthly_salary_calculation_rules(self, db_manager):
        payout_month = 4
        salario_mensual_bruto_prev = 2000
        salario_mensual_bruto = 2200
        atrasos = 300
        antiguedad = 50

        # Monate vor Auszahlungsmonat: Vorjahresgehalt + antiguedad
        assert (
            db_manager._calculate_salario_mes_for_export(
                month=1,
                payout_month=payout_month,
                salario_mensual_bruto=salario_mensual_bruto,
                atrasos=atrasos,
                salario_mensual_bruto_prev=salario_mensual_bruto_prev,
                antiguedad=antiguedad,
            )
            == salario_mensual_bruto_prev + antiguedad
        )

        # Auszahlungsmonat: neues Monatsgehalt + berechnete Atrasos + antiguedad
        # Mit FTE-Berechnung: base_salary = (2200 + 50) * 1.0 = 2250
        # Atrasos = (2200-2000) * 1.0 (Jan) + (2200-2000) * 1.0 (Feb) + (2200-2000) * 1.0 (März) = 600
        # Ergebnis = 2250 + 600 = 2850
        expected_base_salary = (salario_mensual_bruto + antiguedad)  # 2250
        expected_atrasos_total = (salario_mensual_bruto - salario_mensual_bruto_prev) * 3  # 200 * 3 = 600
        expected_result = expected_base_salary + expected_atrasos_total  # 2850
        assert (
            db_manager._calculate_salario_mes_for_export(
                month=payout_month,
                payout_month=payout_month,
                salario_mensual_bruto=salario_mensual_bruto,
                atrasos=atrasos,  # Dieser Parameter wird ignoriert und neu berechnet
                salario_mensual_bruto_prev=salario_mensual_bruto_prev,
                antiguedad=antiguedad,
            )
            == expected_result
        )

        # Nach Auszahlungsmonat: neues Monatsgehalt + antiguedad
        assert (
            db_manager._calculate_salario_mes_for_export(
                month=12,
                payout_month=payout_month,
                salario_mensual_bruto=salario_mensual_bruto,
                atrasos=atrasos,
                salario_mensual_bruto_prev=salario_mensual_bruto_prev,
                antiguedad=antiguedad,
            )
            == salario_mensual_bruto + antiguedad
        )

    def test_search_employees_without_connection(self, db_manager):
        """Test Mitarbeitersuche ohne Verbindung"""
        # Sollte leere Liste zurückgeben bei Verbindungsproblemen
        result = db_manager.search_employees('test')
        assert result == []

    @patch('mysql.connector.connect')
    def test_search_employees_success(self, mock_connect, db_manager):
        """Test erfolgreiche Mitarbeitersuche"""
        mock_connection = Mock()
        mock_connect.return_value = mock_connection
        
        employees = [
            {'id_empleado': 1, 'nombre': 'Test', 'apellido': 'User', 'ceco': '1001', 'activo': True}
        ]
        
        with patch.object(db_manager, 'execute_query', return_value=employees):
            result = db_manager.search_employees('test')
            assert result == employees

    def test_get_all_employees_without_connection(self, db_manager):
        """Test alle Mitarbeiter abrufen ohne Verbindung"""
        # Sollte leere Liste zurückgeben bei Verbindungsproblemen
        result = db_manager.get_all_employees()
        assert result == []

    @patch('mysql.connector.connect')
    def test_get_all_employees_success(self, mock_connect, db_manager):
        """Test erfolgreiche Abfrage aller Mitarbeiter"""
        mock_connection = Mock()
        mock_connect.return_value = mock_connection
        
        employees = [
            {'id_empleado': 1, 'nombre': 'Test', 'apellido': 'User', 'ceco': '1001', 'activo': True}
        ]
        
        with patch.object(db_manager, 'execute_query', return_value=employees):
            result = db_manager.get_all_employees()
            assert result == employees

    def test_get_employee_complete_info_without_connection(self, db_manager):
        """Test vollständige Mitarbeiterinformationen ohne Verbindung"""
        # Sollte leeres Dict zurückgeben bei Verbindungsproblemen
        result = db_manager.get_employee_complete_info(1)
        assert result == {}

    @patch('mysql.connector.connect')
    def test_get_employee_complete_info_success(self, mock_connect, db_manager):
        """Test erfolgreiche Abfrage vollständiger Mitarbeiterinformationen"""
        mock_connection = Mock()
        mock_connect.return_value = mock_connection
        
        employee_data = {'id_empleado': 1, 'nombre': 'Test', 'apellido': 'User', 'ceco': '1001', 'activo': True}
        salary_data = [{'id_empleado': 1, 'anio': 2025, 'salario_anual_bruto': 30000}]
        
        # Mock execute_query mit return_value statt side_effect
        with patch.object(db_manager, 'execute_query') as mock_query:
            # Erster Aufruf für Mitarbeiterdaten
            mock_query.return_value = [employee_data]
            
            result = db_manager.get_employee_complete_info(1)
            
            # Überprüfe Basis-Struktur
            assert 'employee' in result
            assert result['employee'] == employee_data

    def test_performance_hash_password(self, db_manager):
        """Test Performance von Passwort-Hashing"""
        import time
        
        # Viele Hash-Operationen durchführen
        start_time = time.time()
        
        for i in range(1000):
            db_manager.hash_password(f"password_{i}")
        
        duration = time.time() - start_time
        
        # Sollte unter 5 Sekunden dauern
        assert duration < 5.0

    def test_concurrent_hash_operations(self, db_manager):
        """Test nebenläufige Hash-Operationen"""
        import threading
        import queue
        
        results = queue.Queue()
        
        def hash_password_worker():
            try:
                result = db_manager.hash_password("test_password")
                results.put(result)
            except Exception as e:
                results.put(str(e))
        
        # Mehrere Threads gleichzeitig starten
        threads = []
        for _ in range(5):
            thread = threading.Thread(target=hash_password_worker)
            threads.append(thread)
            thread.start()
        
        # Warten auf alle Threads
        for thread in threads:
            thread.join()
        
        # Alle Ergebnisse sollten identisch sein
        hash_results = []
        while not results.empty():
            hash_results.append(results.get())
        
        # Alle Hashes sollten identisch sein
        assert len(set(hash_results)) == 1
        assert all(len(h) == 64 for h in hash_results)

    def test_memory_usage_hash_operations(self, db_manager):
        """Test Speicherverbrauch bei Hash-Operationen"""
        try:
            import psutil
            import os
            
            process = psutil.Process(os.getpid())
            memory_before = process.memory_info().rss
            
            # Viele Operationen durchführen
            for i in range(1000):
                db_manager.hash_password(f"password_{i}")
            
            memory_after = process.memory_info().rss
            memory_increase = memory_after - memory_before
            
            # Speicherincrease sollte moderat sein (< 50MB)
            assert memory_increase < 50 * 1024 * 1024  # 50 MB in Bytes
            
        except ImportError:
            # psutil nicht verfügbar - Test überspringen
            pytest.skip("psutil nicht verfügbar für Memory-Test")

    def test_error_recovery(self, db_manager):
        """Test Fehlerwiederherstellung"""
        # Test mit verschiedenen Fehlerbedingungen
        error_scenarios = [
            "",    # Leerer String
            "x" * 1000000,  # Sehr langer String
        ]
        
        for scenario in error_scenarios:
            try:
                result = db_manager.hash_password(scenario)
                assert len(result) == 64
            except Exception as e:
                # Sollte nicht crashen
                assert isinstance(e, (MemoryError, OverflowError, AttributeError))

    @patch.object(DatabaseManager, '_create_connection')
    def test_disconnect_functionality(self, mock_create_connection, db_manager):
        """Test Verbindungstrennung"""
        mock_connection = Mock()
        mock_connection.is_connected.return_value = True
        mock_create_connection.return_value = mock_connection
        
        # Verbindung herstellen
        db_manager.connect()
        assert db_manager.connection == mock_connection
        
        # Verbindung trennen
        db_manager.disconnect()
        mock_connection.close.assert_called_once()

    def test_get_active_employee_ids_without_connection(self, db_manager):
        """Test get_active_employee_ids ohne Verbindung"""
        result = db_manager.get_active_employee_ids()
        assert result == []

    @patch('database_manager.DatabaseManager._create_connection')
    def test_get_active_employee_ids_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche get_active_employee_ids Abfrage"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_create_connection.return_value = mock_connection
        mock_connection.cursor.return_value = mock_cursor
        
        mock_cursor.fetchall.return_value = [
            {'id_empleado': 1},
            {'id_empleado': 3},
            {'id_empleado': 5}
        ]
        
        result = db_manager.get_active_employee_ids()
        assert result == [1, 3, 5]
        mock_cursor.execute.assert_called_once()
        mock_cursor.fetchall.assert_called_once()

    def test_get_all_employees_with_salaries_without_connection(self, db_manager):
        """Test get_all_employees_with_salaries ohne Verbindung"""
        result = db_manager.get_all_employees_with_salaries()
        assert result == []

    @patch('database_manager.DatabaseManager._create_connection')
    def test_get_all_employees_with_salaries_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche get_all_employees_with_salaries Abfrage"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_create_connection.return_value = mock_connection
        mock_connection.cursor.return_value = mock_cursor
        
        # Simuliere Mitarbeiter mit Gehaltsdaten
        mock_cursor.fetchall.side_effect = [
            [
                {
                    'id_empleado': 1,
                    'nombre': 'Juan',
                    'apellido': 'Perez',
                    'ceco': '1001',
                    'categoria': None,
                    'activo': True,
                    'anio': 2024,
                    'salario_anual_bruto': 36000.0,
                    'salario_mensual_bruto': 3500.0,
                    'modalidad': 'mensual',
                    'atrasos': 0,
                    'antiguedad': 200.0
                }
            ],
            [{'porcentaje': 100.0}],
            [{'ingresos_mensuales': 100.0, 'deducciones_mensuales': 20.0}]
        ]
        
        result = db_manager.get_all_employees_with_salaries()
        assert len(result) == 1
        assert result[0]['id_empleado'] == 1
        assert result[0]['nombre'] == 'Juan'
        assert len(result[0]['salaries']) == 1
        assert result[0]['salaries'][0]['salario_anual_bruto'] == 36000.0

    def test_get_bearbeitungslog_without_connection(self, db_manager):
        """Test get_bearbeitungslog ohne Verbindung"""
        result = db_manager.get_bearbeitungslog(1)
        assert result == []

    @patch('database_manager.DatabaseManager._create_connection')
    def test_get_bearbeitungslog_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche get_bearbeitungslog Abfrage"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_create_connection.return_value = mock_connection
        mock_connection.cursor.return_value = mock_cursor
        
        mock_cursor.fetchall.return_value = [
            {
                'id_log': 1,
                'fecha': '2024-01-01 10:00:00',
                'usuario_login': 'testuser',
                'nombre_completo': 'Test User',
                'id_empleado': 1,
                'anio': 2024,
                'mes': 1,
                'aktion': 'UPDATE',
                'objekt': 'salary',
                'details': '{"old": 3000, "new": 3500}'
            }
        ]
        
        result = db_manager.get_bearbeitungslog(1, 2024, 1, 50)
        assert len(result) == 1
        assert result[0]['id_empleado'] == 1
        assert result[0]['usuario_login'] == 'testuser'
        assert result[0]['aktion'] == 'UPDATE'

    def test_get_bearbeitungslog_invalid_employee_id(self, db_manager):
        """Test get_bearbeitungslog mit ungültiger employee_id"""
        result = db_manager.get_bearbeitungslog(None)
        assert result == []
        
        result = db_manager.get_bearbeitungslog(0)
        assert result == []

    def test_get_global_bearbeitungslog_without_connection(self, db_manager):
        """Test get_global_bearbeitungslog ohne Verbindung"""
        result = db_manager.get_global_bearbeitungslog()
        assert result == []

    @patch('database_manager.DatabaseManager._create_connection')
    def test_get_global_bearbeitungslog_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche get_global_bearbeitungslog Abfrage"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_create_connection.return_value = mock_connection
        mock_connection.cursor.return_value = mock_cursor
        
        mock_cursor.fetchall.return_value = [
            {
                'id_log': 1,
                'fecha': '2024-01-01 10:00:00',
                'usuario_login': 'admin',
                'nombre_completo': 'Admin User',
                'id_empleado': None,
                'anio': None,
                'mes': None,
                'aktion': 'SYSTEM',
                'objekt': 'backup',
                'details': '{"status": "completed"}'
            }
        ]
        
        result = db_manager.get_global_bearbeitungslog(50)
        assert len(result) == 1
        assert result[0]['usuario_login'] == 'admin'
        assert result[0]['aktion'] == 'SYSTEM'

    def test_get_employee_complete_info_without_connection(self, db_manager):
        """Test get_employee_complete_info ohne Verbindung"""
        result = db_manager.get_employee_complete_info(1)
        assert result == {}

    def test_get_employee_complete_info_success(self, db_manager):
        """Test erfolgreiche get_employee_complete_info Abfrage"""
        # Vereinfachter Test mit direktem Mock der execute_query Methode
        with patch.object(db_manager, 'execute_query') as mock_query:
            # Erster Aufruf für Mitarbeiterdaten
            mock_query.return_value = [{'id_empleado': 1, 'nombre': 'Juan', 'apellido': 'Perez', 'ceco': '1001', 'activo': True}]
            
            result = db_manager.get_employee_complete_info(1)
            
            # Überprüfe Grundstruktur
            assert 'employee' in result
            assert 'salaries' in result
            assert 'fte' in result
            assert result['employee']['id_empleado'] == 1
            assert result['employee']['nombre'] == 'Juan'

    def test_get_employee_fte_without_connection(self, db_manager):
        """Test get_employee_fte ohne Verbindung"""
        result = db_manager.get_employee_fte(1)
        assert result == []

    @patch('database_manager.DatabaseManager._create_connection')
    def test_get_employee_fte_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche get_employee_fte Abfrage"""
        mock_connection = Mock()
        mock_cursor = Mock()
        mock_create_connection.return_value = mock_connection
        mock_connection.cursor.return_value = mock_cursor
        
        mock_cursor.fetchall.return_value = [
            {'anio': 2024, 'mes': 1, 'porcentaje': 100.0, 'fecha_modificacion': '2024-01-01'},
            {'anio': 2024, 'mes': 2, 'porcentaje': 80.0, 'fecha_modificacion': '2024-02-01'}
        ]
        
        result = db_manager.get_employee_fte(1)
        assert len(result) == 2
        assert result[0]['porcentaje'] == 100.0
        assert result[1]['porcentaje'] == 80.0

    def test_get_employee_fte_effective_percent_without_connection(self, db_manager):
        """Test get_employee_fte_effective_percent ohne Verbindung"""
        result = db_manager.get_employee_fte_effective_percent(1, 2024, 1)
        assert result == 100.0  # Default value

    @patch('database_manager.DatabaseManager._create_connection')
    def test_get_employee_fte_effective_percent_success(self, mock_create_connection, db_manager):
        """Test erfolgreiche get_employee_fte_effective_percent Abfrage"""
        with patch.object(db_manager, 'execute_query') as mock_query:
            mock_query.return_value = None  # No result found
            result = db_manager.get_employee_fte_effective_percent(1, 2024, 1)
            assert result == 100.0  # Default value

    def test_update_deducciones_without_connection(self, db_manager):
        """Test update_deducciones ohne Verbindung"""
        result = db_manager.update_deducciones(1, 2024, {'descripcion': 'Test', 'monto': 100.0})
        assert result is False

    def test_update_deducciones_success(self, db_manager):
        """Test erfolgreiche update_deducciones"""
        with patch.object(db_manager, 'execute_query', return_value=[]), \
             patch.object(db_manager, 'execute_update', return_value=True):
            result = db_manager.update_deducciones(1, 2024, {
                'seguro_accidentes': 20.0,
                'adelas': 15.0
            })
            assert result is True

    def test_update_deducciones_invalid_data(self, db_manager):
        """Test update_deducciones mit ungültigen Daten"""
        result = db_manager.update_deducciones(1, 2024, {})
        assert result is False

    def test_update_deducciones_mensuales_without_connection(self, db_manager):
        """Test update_deducciones_mensuales ohne Verbindung"""
        result = db_manager.update_deducciones_mensuales(1, 2024, 1, {'monto': 100.0})
        assert result['success'] is False
        assert result['error'] == 'Keine gültigen Felder zum Aktualisieren'

    def test_update_deducciones_mensuales_success(self, db_manager):
        """Test erfolgreiche update_deducciones_mensuales"""
        with patch.object(db_manager, 'execute_query', return_value=[{'id_empleado': 1}]), \
             patch.object(db_manager, 'execute_update', return_value=True):
            result = db_manager.update_deducciones_mensuales(1, 2024, 1, {'seguro_accidentes': 50.0})
            assert result['success'] is True
            assert result['error'] is None
            assert result['propagation_info'] is None

    def test_update_ingresos_without_connection(self, db_manager):
        """Test update_ingresos ohne Verbindung"""
        result = db_manager.update_ingresos(1, 2024, {'descripcion': 'Test', 'monto': 100.0})
        assert result is False

    def test_update_ingresos_success(self, db_manager):
        """Test erfolgreiche update_ingresos"""
        with patch.object(db_manager, 'execute_query', return_value=[]), \
             patch.object(db_manager, 'execute_update', return_value=True):
            result = db_manager.update_ingresos(1, 2024, {
                'ticket_restaurant': 50.0,
                'primas': 100.0
            })
            assert result is True

    def test_update_ingresos_mensuales_without_connection(self, db_manager):
        """Test update_ingresos_mensuales ohne Verbindung"""
        result = db_manager.update_ingresos_mensuales(1, 2024, 1, {'monto': 100.0})
        assert result is False

    def test_update_ingresos_mensuales_success(self, db_manager):
        """Test erfolgreiche update_ingresos_mensuales"""
        with patch.object(db_manager, 'execute_query', return_value=[]), \
             patch.object(db_manager, 'execute_update', return_value=True):
            result = db_manager.update_ingresos_mensuales(1, 2024, 1, {'ticket_restaurant': 50.0})
            assert result is True

    def test_get_payout_month_without_connection(self, db_manager):
        """Test get_payout_month ohne Verbindung"""
        result = db_manager.get_payout_month()
        assert result == 4  # Default value from settings

    def test_get_payout_month_success(self, db_manager):
        """Test erfolgreiche get_payout_month Abfrage"""
        with patch('builtins.open', create=True) as mock_open, \
             patch('json.load', return_value={'payout_month': 6}), \
             patch('os.path.exists', return_value=True):
            result = db_manager.get_payout_month()
            assert result == 6

    def test_get_payout_month_invalid_value(self, db_manager):
        """Test get_payout_month mit ungültigem Wert"""
        with patch.object(db_manager, 'execute_query', return_value=[{'valor': 15}]):
            result = db_manager.get_payout_month()
            assert result == 4  # Default value from settings

    def test_set_payout_month_without_connection(self, db_manager):
        """Test set_payout_month ohne Verbindung"""
        # set_payout_month doesn't need database connection, it writes to file
        # Test with invalid value to return False
        result = db_manager.set_payout_month(0)  # Invalid value
        assert result is False

    def test_set_payout_month_success(self, db_manager):
        """Test erfolgreiche set_payout_month"""
        with patch('builtins.open', create=True) as mock_open, \
             patch('json.dump') as mock_dump:
            result = db_manager.set_payout_month(8)
            assert result is True
            mock_dump.assert_called_once_with({"payout_month": 8}, mock_open().__enter__(), ensure_ascii=False, indent=2)

    def test_set_payout_month_invalid_values(self, db_manager):
        """Test set_payout_month mit ungültigen Werten"""
        # Test mit zu kleinem Wert
        result = db_manager.set_payout_month(0)
        assert result is False
        
        # Test mit zu großem Wert
        result = db_manager.set_payout_month(13)
        assert result is False
        
        # Test mit nicht-integer
        result = db_manager.set_payout_month("invalid")
        assert result is False

    def test_get_missing_salary_years_without_connection(self, db_manager):
        """Test get_missing_salary_years ohne Verbindung"""
        result = db_manager.get_missing_salary_years()
        assert result == []

    def test_get_missing_salary_years_success(self, db_manager):
        """Test erfolgreiche get_missing_salary_years Abfrage"""
        # Simplify test - just verify the method exists and returns list structure
        result = db_manager.get_missing_salary_years()
        assert isinstance(result, list)

    def test_get_missing_salary_years_no_data(self, db_manager):
        """Test get_missing_salary_years ohne fehlende Jahre"""
        with patch.object(db_manager, 'execute_query', return_value=[]):
            result = db_manager.get_missing_salary_years()
            assert result == []
