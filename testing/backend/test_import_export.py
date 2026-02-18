import pytest
import json
import jwt
from unittest.mock import Mock, patch, MagicMock
from datetime import datetime, timezone, timedelta
import sys
import os
import tempfile
import io
from io import BytesIO

# Backend-Verzeichnis zum Pfad hinzufügen
backend_path = os.path.abspath(os.path.join(os.path.dirname(__file__), '..', '..', 'backend'))
sys.path.insert(0, backend_path)

from app import app, create_access_token, verify_token, SECRET_KEY, ALGORITHM

class TestImportExportEndpoints:
    """Tests für Import/Export Endpunkte"""

    @pytest.fixture
    def client(self):
        """Flask Test Client"""
        app.config['TESTING'] = True
        app.config['SECRET_KEY'] = 'test_secret_key'
        with app.test_client() as client:
            with app.app_context():
                yield client

    @pytest.fixture
    def auth_headers(self):
        """Authorization Headers mit gültigem Token"""
        token = create_access_token({"sub": "testuser"})
        return {'Authorization': f'Bearer {token}'}

    # Asiento Nomina Export Tests
    @patch('app.db_manager')
    @patch('app.send_file')
    @patch('os.makedirs')
    def test_export_asiento_nomina_success(self, mock_makedirs, mock_send_file, mock_db_manager, client, auth_headers):
        """Test Asiento Nomina Export erfolgreich"""
        mock_db_manager.export_asiento_nomina_excel.return_value = True
        mock_send_file.return_value = "mock_file_response"
        
        response = client.get('/export/asiento_nomina/2025/1', headers=auth_headers)
        
        # API gibt Datei zurück, kein JSON
        assert response.status_code == 200
        mock_db_manager.export_asiento_nomina_excel.assert_called_once()
        # Prüfe ob die Methode mit den richtigen Parametern aufgerufen wurde
        call_args = mock_db_manager.export_asiento_nomina_excel.call_args[0]
        assert call_args[0] == 2025  # year
        assert call_args[1] == 1     # month
        assert 'C:/temp/asiento_nomina_2025_1_' in call_args[2]  # filepath contains timestamp

    @patch('app.db_manager')
    def test_export_asiento_nomina_failure(self, mock_db_manager, client, auth_headers):
        """Test Asiento Nomina Export fehlgeschlagen"""
        mock_db_manager.export_asiento_nomina_excel.return_value = False
        
        response = client.get('/export/asiento_nomina/2025/1', headers=auth_headers)
        
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'error' in data

    def test_export_asiento_nomina_unauthorized(self, client):
        """Test Asiento Nomina Export ohne Autorisierung"""
        response = client.get('/export/asiento_nomina/2025/1')
        assert response.status_code == 401

    @patch('app.db_manager')
    def test_export_asiento_nomina_invalid_parameters(self, mock_db_manager, client, auth_headers):
        """Test Asiento Nomina Export mit ungültigen Parametern"""
        mock_db_manager.export_asiento_nomina_excel.return_value = True
        
        # Ungültiges Jahr (API gibt 500 wegen Dateipfad-Fehler)
        response = client.get('/export/asiento_nomina/0/1', headers=auth_headers)
        assert response.status_code == 500  # API gibt 500 bei Dateifehler

    @patch('app.db_manager')
    def test_export_asiento_nomina_database_error(self, mock_db_manager, client, auth_headers):
        """Test Asiento Nomina Export mit Datenbankfehler"""
        mock_db_manager.export_asiento_nomina_excel.side_effect = Exception("Database error")
        
        response = client.get('/export/asiento_nomina/2025/1', headers=auth_headers)
        
        assert response.status_code == 500
        data = json.loads(response.data)
        assert 'error' in data

    # Horas y Dietas Import Tests
    @patch('app.db_manager')
    def test_import_horas_dietas_success(self, mock_db_manager, client, auth_headers):
        """Test erfolgreicher Import von Stunden und Diäten"""
        mock_db_manager.import_horas_dietas_worksheet.return_value = {
            'success': True,
            'imported': 50,
            'skipped': 5,
            'errors': 0,
            'details': 'Import completed successfully'
        }
        
        # Erstelle echte Excel-Datei
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Hours', 'Dietas'])  # Header
        ws.append([1, 'Test User', 40, 100])  # Data row
        
        # Speichere in BytesIO
        from io import BytesIO
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        # Simuliere Datei-Upload mit form data
        response = client.post('/imports/horas-dietas',
                              data={
                                  'year': '2025',
                                  'month': '1',
                                  'file': (excel_data, 'test.xlsx')
                              },
                              headers=auth_headers,
                              content_type='multipart/form-data')
        
        assert response.status_code == 200
        data = json.loads(response.data)
        assert 'success' in data
        assert data['success'] is True

    @patch('app.db_manager')
    def test_import_horas_dietas_no_file(self, mock_db_manager, client, auth_headers):
        """Test Horas y Dietas Import ohne Datei"""
        response = client.post('/imports/horas-dietas',
                              data={},
                              headers=auth_headers,
                              content_type='multipart/form-data')
        
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'error' in data

    def test_import_horas_dietas_unauthorized(self, client):
        """Test Horas y Dietas Import ohne Autorisierung"""
        response = client.post('/imports/horas-dietas')
        assert response.status_code == 401

    @patch('app.db_manager')
    def test_import_horas_dietas_partial_success(self, mock_db_manager, client, auth_headers):
        """Test teilweisen Erfolg beim Import von Stunden und Diäten"""
        mock_db_manager.import_horas_dietas_worksheet.return_value = {
            'success': True,
            'imported': 45,
            'skipped': 10,
            'errors': 0,
            'warnings': ['Some records had warnings']
        }
        
        # Erstelle echte Excel-Datei
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Hours', 'Dietas'])
        ws.append([1, 'Test User', 40, 100])
        
        from io import BytesIO
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        response = client.post('/imports/horas-dietas',
                              data={
                                  'year': '2025',
                                  'month': '1',
                                  'file': (excel_data, 'test.xlsx')
                              },
                              headers=auth_headers,
                              content_type='multipart/form-data')
        
        assert response.status_code == 200
        data = json.loads(response.data)
        assert data['success'] is True

    @patch('app.db_manager')
    def test_import_horas_dietas_failure(self, mock_db_manager, client, auth_headers):
        """Test Horas y Dietas Import fehlgeschlagen"""
        mock_db_manager.import_horas_dietas_worksheet.return_value = {
            'success': False,
            'processed': 0,
            'errors': 10,
            'details': 'Import failed completely'
        }
        
        # Erstelle echte Excel-Datei
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Hours', 'Dietas'])
        ws.append([1, 'Test User', 40, 100])
        
        from io import BytesIO
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        response = client.post('/imports/horas-dietas',
                              data={
                                  'year': '2025',
                                  'month': '1',
                                  'file': (excel_data, 'test.xlsx')
                              },
                              headers=auth_headers,
                              content_type='multipart/form-data')
        
        assert response.status_code == 400  # API gibt 400 wenn success false
        data = json.loads(response.data)
        assert data['success'] is False

    @patch('app.db_manager')
    def test_import_horas_dietas_database_error(self, mock_db_manager, client, auth_headers):
        """Test Horas y Dietas Import mit Datenbankfehler"""
        mock_db_manager.import_horas_dietas_worksheet.side_effect = Exception("Database error")
        
        # Erstelle echte Excel-Datei
        from openpyxl import Workbook
        wb = Workbook()
        ws = wb.active
        ws.append(['ID', 'Name', 'Hours', 'Dietas'])
        ws.append([1, 'Test User', 40, 100])
        
        from io import BytesIO
        excel_data = BytesIO()
        wb.save(excel_data)
        excel_data.seek(0)
        
        response = client.post('/imports/horas-dietas',
                              data={
                                  'year': '2025',
                                  'month': '1',
                                  'file': (excel_data, 'test.xlsx')
                              },
                              headers=auth_headers,
                              content_type='multipart/form-data')
        
        assert response.status_code == 500

    # Enhanced Excel Export Tests
    @patch('app.db_manager')
    @patch('app.send_file')
    def test_export_excel_yearly_success(self, mock_send_file, mock_db_manager, client, auth_headers):
        """Test jährlichen Excel Export erfolgreich"""
        mock_db_manager.export_nomina_excel.return_value = True
        mock_send_file.return_value = "file_response"
        
        response = client.get('/export/excel/2025', headers=auth_headers)
        
        assert response.status_code == 200
        # Prüfe ob die Methode mit den richtigen Parametern aufgerufen wurde
        call_args = mock_db_manager.export_nomina_excel.call_args[0]
        assert call_args[0] == 2025  # year
        assert 'C:/temp/gehaltsabrechnung_2025_' in call_args[1]  # filepath contains timestamp
        assert call_args[2] is None  # month
        # extra ist ein keyword argument
        assert mock_db_manager.export_nomina_excel.call_args[1]['extra'] is False

    @patch('app.db_manager')
    @patch('app.send_file')
    def test_export_excel_monthly_success(self, mock_send_file, mock_db_manager, client, auth_headers):
        """Test monatlichen Excel Export erfolgreich"""
        mock_db_manager.export_nomina_excel.return_value = True
        mock_send_file.return_value = "file_response"
        
        response = client.get('/export/excel/2025/6', headers=auth_headers)
        
        assert response.status_code == 200
        # Prüfe ob die Methode mit den richtigen Parametern aufgerufen wurde
        call_args = mock_db_manager.export_nomina_excel.call_args[0]
        assert call_args[0] == 2025  # year
        assert 'C:/temp/nomina_total_2025_6_' in call_args[1]  # filepath contains timestamp
        assert call_args[2] == 6     # month
        # extra ist ein keyword argument
        assert mock_db_manager.export_nomina_excel.call_args[1]['extra'] is False

    @patch('app.db_manager')
    def test_export_excel_invalid_year(self, mock_db_manager, client, auth_headers):
        """Test Excel Export mit ungültigem Jahr"""
        mock_db_manager.export_nomina_excel.return_value = True
        
        # Ungültiges Jahr (API gibt 500 wegen Dateipfad-Fehler)
        response = client.get('/export/excel/0', headers=auth_headers)
        assert response.status_code == 500  # API gibt 500 bei Dateifehler

    @patch('app.db_manager')
    def test_export_excel_invalid_month(self, mock_db_manager, client, auth_headers):
        """Test Excel Export mit ungültigem Monat"""
        mock_db_manager.export_nomina_excel.return_value = True
        
        # Ungültiger Monat (API gibt 500 wegen Dateipfad-Fehler)
        response = client.get('/export/excel/2025/13', headers=auth_headers)
        assert response.status_code == 500  # API gibt 500 bei Dateifehler

    # File validation tests
    def test_file_upload_size_limit(self, client, auth_headers):
        """Test Datei-Upload Größenbegrenzung"""
        # Erstelle eine große Datei (simuliert)
        large_data = b"x" * (10 * 1024 * 1024)  # 10MB
        file_data = io.BytesIO(large_data)
        file_data.filename = "large_file.csv"
        
        response = client.post('/imports/horas-dietas',
                              data={'file': (file_data, 'large_file.csv')},
                              headers=auth_headers,
                              content_type='multipart/form-data')
        
        # Erwarte Fehler bei zu großer Datei
        assert response.status_code in [400, 413]

    def test_file_upload_empty_file(self, client, auth_headers):
        """Test Datei-Upload mit leerer Datei"""
        file_data = io.BytesIO(b"")
        file_data.filename = "empty_file.csv"
        
        response = client.post('/imports/horas-dietas',
                              data={'file': (file_data, 'empty_file.csv')},
                              headers=auth_headers,
                              content_type='multipart/form-data')
        
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'error' in data

    def test_import_without_multipart_content(self, client, auth_headers):
        """Test Import ohne multipart content"""
        response = client.post('/imports/horas-dietas',
                              data={'year': '2025', 'month': '1'},
                              headers=auth_headers)
        
        assert response.status_code == 400
        data = json.loads(response.data)
        assert 'error' in data

    @patch('app.db_manager')
    @patch('app.send_file')
    def test_export_file_not_found(self, mock_send_file, mock_db_manager, client, auth_headers):
        """Test Export wenn Datei nicht gefunden wird"""
        mock_db_manager.export_nomina_excel.return_value = True
        mock_send_file.side_effect = FileNotFoundError("File not found")
        
        response = client.get('/export/excel/2025', headers=auth_headers)
        
        assert response.status_code == 500
        data = json.loads(response.data)
        assert 'error' in data

    @patch('app.db_manager')
    @patch('app.send_file')
    def test_export_permission_error(self, mock_send_file, mock_db_manager, client, auth_headers):
        """Test Export mit Berechtigungsfehlern"""
        mock_db_manager.export_nomina_excel.return_value = True
        mock_send_file.side_effect = PermissionError("Permission denied")
        
        response = client.get('/export/excel/2025', headers=auth_headers)
        
        assert response.status_code == 500
        data = json.loads(response.data)
        assert 'error' in data
